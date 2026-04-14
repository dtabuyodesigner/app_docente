# ==============================================================================
# ROUTES/INFORMES.PY - GESTIÓN DE INFORMES Y PDFS
# ==============================================================================
from flask import Blueprint, jsonify, request, send_file, session
from utils.db import get_db, nivel_a_nota
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from io import BytesIO
import json
import pandas as pd
from datetime import date, datetime
import os
import re

# ==============================================================================
# AUXILIARES PARA INFORMES
# ==============================================================================

def get_logo_path():
    """Busca el logotipo de la aplicación."""
    possible_paths = [
        os.path.join(os.getcwd(), "static", "icon-512.png"),
        os.path.join(os.path.dirname(__file__), "..", "static", "icon-512.png"),
        "/opt/cuaderno-del-tutor/static/icon-512.png"
    ]
    for p in possible_paths:
        if os.path.exists(p):
            return p
    return None

def add_header(elements, styles, titulo, colegio="", fecha=""):
    """Añade una cabecera con el nombre del centro, título y fecha."""
    header_table_data = [
        [Paragraph(f"Centro: {colegio}" if colegio else " ", styles['Normal']),
         Paragraph(f"Fecha: {fecha}" if fecha else " ", styles['Normal'])]
    ]
    header_table = Table(header_table_data, colWidths=[10*cm, 6*cm])
    header_table.setStyle(TableStyle([
        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 10))
    elements.append(Paragraph(titulo, styles['Title']))
    elements.append(Spacer(1, 12))

def generar_grafica_rendimiento(notas_area, alumno_nombre, es_infantil):
    """Genera la gráfica de barras de rendimiento."""
    if not notas_area:
        return None
    
    # Filtrar áreas que no tienen ninguna evaluación para evitar errores con NoneType
    datos_validos = [row for row in notas_area if row[2] is not None]
    if not datos_validos:
        return None
        
    areas = [row[1] for row in datos_validos]
    notas = [row[2] for row in datos_validos]
    escalas = [row[3] for row in datos_validos]

    fig, ax = plt.subplots(figsize=(8, max(3, len(areas) * 0.5)))

    colors_bars = []
    for n, es in zip(notas, escalas):
        if (es and es.startswith("INFANTIL_")):
            if n < 1.5:
                colors_bars.append('#dc3545')
            elif n < 2.5:
                colors_bars.append('#ffc107')
            else:
                colors_bars.append('#28a745')
        else:
            colors_bars.append('#28a745' if n >= 5 else '#dc3545')
        
    bars = ax.barh(areas, notas, color=colors_bars)
    ax.set_xlabel('Nota Media')
    ax.set_title(f'Rendimiento por Área - {alumno_nombre}')
    has_num = any((not es or not es.startswith("INFANTIL_")) for es in escalas)
    ax.set_xlim(0, 10 if has_num else 3.5)
    ax.grid(axis='x', alpha=0.3)

    for i, (bar, nota, es) in enumerate(zip(bars, notas, escalas)):
        def local_format_nota(n, e):
            if n is None:
                return "—"
            if (e and e.startswith("INFANTIL_")):
                rnd = round(n)
                if e == "INFANTIL_PA_A_MA":
                    return {1: "PA", 2: "AD", 3: "MA"}.get(rnd, "—")
                return {1: "No Iniciado", 2: "En proceso", 3: "Conseguido"}.get(rnd, "—")
            return f"{n:.2f}"
        ax.text(nota + 0.2, i, local_format_nota(nota, es), va='center', fontweight='bold')

    chart_buf = BytesIO()
    plt.savefig(chart_buf, format='png', bbox_inches='tight', dpi=300)
    chart_buf.seek(0)
    plt.close()
    return chart_buf

def generar_pie_circular(valores, etiquetas, titulo):
    """Genera un gráfico circular."""
    if not valores or sum(valores) == 0:
        return None
    
    plt.figure(figsize=(4, 3))
    plt.pie(valores, labels=etiquetas, autopct='%1.1f%%', startangle=140)
    plt.title(titulo)
    plt.axis('equal') 

    img_buffer = BytesIO()
    plt.savefig(img_buffer, format='png', bbox_inches='tight', dpi=300)
    plt.close()
    img_buffer.seek(0)
    return img_buffer

def format_nota(nota, tipo_escala):
    """Convierte nota numérica a texto legible según escala."""
    if nota is None:
        return "—"
    if (tipo_escala and tipo_escala.startswith("INFANTIL_")):
        n = round(nota) if nota <= 3 else 0
        if n == 0:
            if nota <= 3.5:
                n = 1
            elif nota <= 6.5:
                n = 2
            else:
                n = 3
        if tipo_escala == "INFANTIL_PA_A_MA":
            return {1: "PA", 2: "AD", 3: "MA"}.get(n, "—")
        return {1: "NI", 2: "EP", 3: "C"}.get(n, "—")
    return f"{nota:.2f}"

# ==============================================================================
# BLUEPRINT
# ==============================================================================

informes_bp = Blueprint('informes', __name__)

# ==============================================================================
# INFORME INDIVIDUAL PDF
# ==============================================================================

@informes_bp.route("/api/informe/pdf_individual")
def informe_pdf_individual():
    alumno_id = request.args.get("alumno_id")
    trimestre = request.args.get("trimestre")
    tutor_nombre = request.args.get("tutor", "")
    area_id_filtro = request.args.get("area_id", "")
    rol = request.args.get("rol", "tutor")
    
    if not alumno_id or not trimestre:
        return "Faltan parámetros", 400
    
    conn = get_db()
    cur = conn.cursor()

    cur.execute("SELECT nombre FROM alumnos WHERE id = ?", (alumno_id,))
    row_al = cur.fetchone()
    if not row_al:
        return "Alumno no encontrado", 404
    alumno = row_al["nombre"]

    area_nombre_filtro = ""
    if rol == "especialista" and area_id_filtro:
        cur.execute("SELECT nombre FROM areas WHERE id = ?", (area_id_filtro,))
        row_area = cur.fetchone()
        if row_area:
            area_nombre_filtro = row_area["nombre"]

    cur.execute("SELECT grupo_id FROM alumnos WHERE id = ?", (alumno_id,))
    row_aid = cur.fetchone()
    grupo_id = row_aid["grupo_id"] if row_aid else None

    row_inf = None
    if grupo_id:
        cur.execute("SELECT * FROM informe_grupo WHERE grupo_id = ? AND trimestre = ?", (grupo_id, trimestre))
        row_inf = cur.fetchone()

    cur.execute("""
        SELECT estado, COUNT(*) 
        FROM asistencia 
        WHERE alumno_id = ?
        GROUP BY estado
    """, (alumno_id,))
    asist_data = dict(cur.fetchall())

    alumnos = [{"id": alumno_id, "nombre": alumno}]

    periodo = f"T{trimestre}"
    area_filter_sql = "AND a.id = ?" if area_id_filtro else ""
    area_filter_params = [area_id_filtro] if area_id_filtro else []

    cur.execute(f"""
        SELECT a.id, a.nombre, ROUND(AVG(val.nota), 2) as media, a.tipo_escala
        FROM (
            SELECT area_id, nota FROM evaluaciones WHERE alumno_id = ? AND trimestre = ?
            UNION ALL
            SELECT c.area_id, ec.nota 
            FROM evaluacion_criterios ec
            JOIN criterios c ON ec.criterio_id = c.id
            WHERE ec.alumno_id = ? AND ec.periodo = ?
        ) val
        JOIN areas a ON val.area_id = a.id
        WHERE 1=1 {area_filter_sql}
        GROUP BY a.id, a.nombre, a.tipo_escala
    """, (alumno_id, trimestre, alumno_id, periodo) + tuple(area_filter_params))
    notas_area = cur.fetchall()

    cur.execute("""
        SELECT a.nombre as area, a.id as area_id, s.nombre as sda, c.codigo as criterio_codigo, 
               c.descripcion as criterio_desc, e.nota, a.tipo_escala, e.nivel, c.comentario_base
        FROM evaluaciones e
        JOIN sda s ON e.sda_id = s.id
        JOIN areas a ON e.area_id = a.id
        JOIN criterios c ON e.criterio_id = c.id
        WHERE e.alumno_id = ? AND e.trimestre = ?
        
        UNION ALL
        
        SELECT a.nombre as area, a.id as area_id, 'Criterios Directos' as sda, c.codigo as criterio_codigo,
               c.descripcion as criterio_desc, ec.nota, a.tipo_escala, ec.nivel, c.comentario_base
        FROM evaluacion_criterios ec
        JOIN criterios c ON ec.criterio_id = c.id
        JOIN areas a ON c.area_id = a.id
        WHERE ec.alumno_id = ? AND ec.periodo = ?
        
        ORDER BY area, sda, criterio_codigo
    """, (alumno_id, trimestre, alumno_id, periodo))
    all_criterios = cur.fetchall()
    
    if area_id_filtro:
        notas_criterios = [r for r in all_criterios if str(r["area_id"]) == str(area_id_filtro)]
    else:
        notas_criterios = list(all_criterios)

    year = date.today().year
    if trimestre == "1":
        start_date, end_date = f"{year-1}-09-01", f"{year-1}-12-31"
    elif trimestre == "2":
        start_date, end_date = f"{year}-01-01", f"{year}-03-31"
    else:
        start_date, end_date = f"{year}-04-01", f"{year}-06-30"

    cur.execute("""
        SELECT fecha, estado
        FROM asistencia
        WHERE alumno_id = ? AND estado IN ('retraso', 'falta_justificada', 'falta_no_justificada')
        AND fecha BETWEEN ? AND ?
        ORDER BY fecha
    """, (alumno_id, start_date, end_date))
    faltas_detalle = cur.fetchall()

    cur.execute("SELECT texto FROM informe_individual WHERE alumno_id = ? AND trimestre = ?", (alumno_id, trimestre))
    obs_row = cur.fetchone()
    obs_text = obs_row["texto"] if obs_row and obs_row["texto"] else ""

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()

    add_header(elements, styles, f"Informe Trimestral - Trimestre {trimestre}")
    elements.append(Paragraph(f"Alumno: {alumno}", styles['Heading2']))
    if tutor_nombre:
        rol_label = f"Especialista de {area_nombre_filtro}" if rol == "especialista" and area_nombre_filtro else "Especialista" if rol == "especialista" else "Tutor/a"
        elements.append(Paragraph(f"{rol_label}: {tutor_nombre}", styles['Normal']))
    elements.append(Spacer(1, 12))

    elements.append(Paragraph("Resumen de Asistencia", styles['Heading3']))
    data_asist = [
        ["Estado", "Días"],
        [Paragraph('<font color="#ffc107">●</font> Retraso', styles['Normal']), asist_data.get('retraso', 0)],
        [Paragraph('<font color="#17a2b8">●</font> Falta Justificada', styles['Normal']), asist_data.get('falta_justificada', 0)],
        [Paragraph('<font color="#dc3545">●</font> Falta No Justificada', styles['Normal']), asist_data.get('falta_no_justificada', 0)],
    ]
    t_asist = Table(data_asist, colWidths=[6*cm, 2*cm])
    t_asist.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE')
    ]))
    elements.append(t_asist)

    if faltas_detalle:
        elements.append(Spacer(1, 8))
        fechas_str = []
        for row in faltas_detalle:
            f = row["fecha"]
            e = row["estado"]
            try:
                d_obj = datetime.strptime(f, '%Y-%m-%d')
                d_str = d_obj.strftime('%d/%m')
            except:
                d_str = f

            if e == 'retraso':
                color, tipo = "#ffc107", "R"
            elif e == 'falta_justificada':
                color, tipo = "#17a2b8", "F"
            else:
                color, tipo = "#dc3545", "F"

            fechas_str.append(f'<font color="{color}">{d_str}({tipo})</font>')
        
        elements.append(Paragraph("<b>Detalle (Día/Tipo):</b> " + ", ".join(fechas_str), styles['Normal']))

    elements.append(Spacer(1, 12))
    elements.append(Paragraph("Rendimiento Académico", styles['Heading3']))

    area_sda_map = {}
    for area, area_id_r, sda, crit_cod, crit_desc, nota, escala, nivel, base in notas_criterios:
        if area not in area_sda_map:
            area_sda_map[area] = {}
        if sda not in area_sda_map[area]:
            area_sda_map[area][sda] = []
        area_sda_map[area][sda].append((f"{crit_cod}", crit_desc or "", nota, escala))
    
    data_notas = [["Área", "Nota Media", "Detalle SDA y Criterios"]]
    for aid, area_nombre, nota_media, tipo_escala in notas_area:
        sda_dict = area_sda_map.get(area_nombre, {})
        detail_lines = []
        for sda_name, criterios in sda_dict.items():
            eval_notas = [c[2] for c in criterios if c[2] is not None]
            sda_avg = round(sum(eval_notas) / len(eval_notas), 2) if eval_notas else 0
            fmt_sda_avg = format_nota(sda_avg, tipo_escala)
            detail_lines.append(f"<b>{sda_name}</b>: {fmt_sda_avg}")
            for crit_cod, crit_desc, nota, escala in criterios:
                desc_txt = f" – {crit_desc}" if crit_desc else ""
                detail_lines.append(f"  • {crit_cod}{desc_txt}: {format_nota(nota, escala)}")
        
        detail_str = "<br/>".join(detail_lines) if detail_lines else "Sin datos"
        data_notas.append([
            Paragraph(area_nombre, styles['Normal']), 
            str(format_nota(nota_media, tipo_escala)), 
            Paragraph(detail_str, styles['Normal'])
        ])
    
    t_notas = Table(data_notas, colWidths=[4*cm, 2.5*cm, 10*cm])
    t_notas.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold')
    ]))
    elements.append(t_notas)
    elements.append(Spacer(1, 15))

    from reportlab.platypus import Image as RLImage
    cur.execute("SELECT tipo_evaluacion FROM grupos WHERE id = ?", (grupo_id,))
    tipo_g = cur.fetchone()
    es_infantil = tipo_g and tipo_g["tipo_evaluacion"] == "infantil"
    chart_buf = generar_grafica_rendimiento(notas_area, alumno, es_infantil)
    if chart_buf:
        elements.append(RLImage(chart_buf, width=16*cm, height=max(4, len(notas_area) * 0.7)*cm))
        elements.append(Spacer(1, 15))

    if obs_text:
        elements.append(Paragraph("Observaciones del Tutor/a:", styles['Heading3']))
        elements.append(Paragraph(obs_text.replace('\n', '<br/>'), styles['BodyText']))
        elements.append(Spacer(1, 15))

    elements.append(Paragraph("Observaciones Pedagógicas", styles['Heading3']))
    comentarios_por_area = {}
    for area, area_id_r, sda, crit_cod, crit_desc, nota, escala, nivel, base in notas_criterios:
        if nivel:
            comment = ""
            if (escala and escala.startswith("INFANTIL_")):
                # Lógica para 3 niveles en Infantil
                is_pama = (escala == "INFANTIL_PA_A_MA")
                if nivel == 1:
                    comment = f"• <b>{crit_cod}</b>: En proceso inicial de adquirir {crit_desc}." if is_pama else f"• <b>{crit_cod}</b>: No ha iniciado {crit_desc}."
                elif nivel == 2:
                    comment = f"• <b>{crit_cod}</b>: Ha progresado adecuadamente en {crit_desc}." if is_pama else f"• <b>{crit_cod}</b>: Está en proceso de adquirir {crit_desc}."
                elif nivel >= 3:
                    comment = f"• <b>{crit_cod}</b>: Ha adquirido plenamente {crit_desc}." if is_pama else f"• <b>{crit_cod}</b>: Ha conseguido {crit_desc}."
            else:
                # Primaria / 4 niveles
                if nivel == 1:
                    comment = f"• <b>{crit_cod}</b>: Necesita apoyo en {crit_desc}."
                elif nivel == 2:
                    comment = f"• <b>{crit_cod}</b>: Está en proceso de mejorar en {crit_desc}."
                elif nivel == 3:
                    comment = f"• <b>{crit_cod}</b>: Comprende y aplica adecuadamente {crit_desc}."
                elif nivel >= 4:
                    comment = f"• <b>{crit_cod}</b>: Destaca especialmente en {crit_desc}."
            
            if base:
                comment += f" {base}"
            
            if area not in comentarios_por_area:
                comentarios_por_area[area] = []
            comentarios_por_area[area].append(comment)
        
    if comentarios_por_area:
        for area, comentarios in comentarios_por_area.items():
            elements.append(Paragraph(f"<b><u>{area}</u></b>", styles['Normal']))
            for c in comentarios:
                elements.append(Paragraph(c, styles['Normal']))
            elements.append(Spacer(1, 4))
    else:
        elements.append(Paragraph("Sin observaciones registradas.", styles['Normal']))

    elements.append(Spacer(1, 40))
    if tutor_nombre:
        rol_label = f"Especialista de {area_nombre_filtro}" if rol == "especialista" and area_nombre_filtro else "Especialista" if rol == "especialista" else "Tutor/a"
        firma_data = [
            [Paragraph(f"Fdo: {tutor_nombre}", styles['Normal'])],
            [Paragraph(f"({rol_label})", styles['Normal'])]
        ]
        t_firma = Table(firma_data, colWidths=[10*cm])
        t_firma.setStyle(TableStyle([
            ('LEFTPADDING', (0,0), (-1,-1), 0),
            ('TOPPADDING', (0,0), (-1,-1), 0),
            ('BOTTOMPADDING', (0,0), (-1,-1), 2),
        ]))
        elements.append(t_firma)

    doc.build(elements)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"Informe_{alumno}_T{trimestre}.pdf",
        mimetype='application/pdf'
    )

# ==============================================================================
# INFORME PDF TODOS - CORREGIDO (ERROR DE BINDINGS)
# ==============================================================================

@informes_bp.route("/api/informe/pdf_general")
def informe_pdf_todos():
    """Genera informe PDF general del grupo - CORREGIDO PARA INFANTIL"""
    trimestre = request.args.get("trimestre")
    tutor_nombre = request.args.get("tutor", "")
    colegio = request.args.get("colegio", "")
    fecha_inf = request.args.get("fecha_informe", "")
    area_id = request.args.get("area_id")
    
    conn = get_db()
    cur = conn.cursor()
    grupo_id = session.get('active_group_id')

    cargo_tutor = "Tutor/a"
    if grupo_id:
        cur.execute(f"SELECT valor FROM config WHERE clave = 'grupo_{grupo_id}_rol'")
        rol_row = cur.fetchone()
        if rol_row and rol_row["valor"] == "especialista":
            cargo_tutor = "Especialista"
            if area_id:
                cur.execute("SELECT nombre FROM areas WHERE id = ?", (area_id,))
                area_row = cur.fetchone()
                if area_row:
                    cargo_tutor = f"Especialista de {area_row['nombre']}"

    cur.execute("SELECT nombre FROM grupos WHERE id = ?", (grupo_id,))
    row_g = cur.fetchone()
    grupo_nombre = row_g["nombre"] if row_g else "Grupo"

    cur.execute("SELECT * FROM informe_grupo WHERE trimestre = ? AND grupo_id = ?", (trimestre, grupo_id))
    row_grupo = cur.fetchone()

    cur.execute("""
        SELECT a.tipo_escala 
        FROM grupos g
        JOIN areas a ON g.etapa_id = a.etapa_id
        WHERE g.id = ? AND a.tipo_escala LIKE 'INFANTIL_%'
        LIMIT 1
    """, (grupo_id,))
    es_infantil = cur.fetchone() is not None

    query_prom = """
        SELECT a.nombre, COUNT(val.area_id) as num_suspensos
        FROM alumnos a
        LEFT JOIN (
            SELECT alumno_id, area_id, AVG(nota) as media_area
            FROM evaluaciones
            WHERE trimestre = ?
    """
    params_prom = [trimestre]
    if area_id:
        query_prom += " AND area_id = ?"
        params_prom.append(area_id)
    
    query_prom += """
            GROUP BY alumno_id, area_id
            HAVING media_area < 5
        ) val ON a.id = val.alumno_id
        WHERE a.grupo_id = ?
        GROUP BY a.id, a.nombre
        HAVING num_suspensos > 0
    """
    params_prom.append(grupo_id)

    cur.execute(query_prom, params_prom)
    suspensos_data = cur.fetchall()

    susp_map = {0: 0, 1: 0, 2: 0, 3: 0}
    for r in suspensos_data:
        n = r["num_suspensos"]
        if n > 2:
            susp_map[3] += 1
        else:
            susp_map[n] += 1
        
    cur.execute("SELECT count(*) FROM alumnos WHERE grupo_id = ?", (grupo_id,))
    total_alumnos = cur.fetchone()[0]
    susp_map[0] = total_alumnos - len(suspensos_data)

    year = date.today().year
    if trimestre == "1":
        start_date, end_date = f"{year-1}-09-01", f"{year-1}-12-31"
    elif trimestre == "2":
        start_date, end_date = f"{year}-01-01", f"{year}-03-31"
    else:
        start_date, end_date = f"{year}-04-01", f"{year}-06-30"

    cur.execute("""
        SELECT a.nombre, asi.estado, COUNT(*) as c
        FROM asistencia asi
        JOIN alumnos a ON a.id = asi.alumno_id
        WHERE asi.fecha BETWEEN ? AND ? AND a.grupo_id = ?
        GROUP BY a.id, a.nombre, asi.estado
        ORDER BY c DESC
    """, (start_date, end_date, grupo_id))
    asist_rows = cur.fetchall()

    max_faltas = {"nombre": "Nadie", "num": 0}
    max_retrasos = {"nombre": "Nadie", "num": 0}

    temp_faltas = {}
    temp_retrasos = {}

    for r in asist_rows:
        if r["estado"] in ["falta_justificada", "falta_no_justificada"]:
            temp_faltas[r["nombre"]] = temp_faltas.get(r["nombre"], 0) + r["c"]
        elif r["estado"] == "retraso":
            temp_retrasos[r["nombre"]] = temp_retrasos.get(r["nombre"], 0) + r["c"]
        
    if temp_faltas:
        u_max = max(temp_faltas, key=temp_faltas.get)
        max_faltas = {"nombre": u_max, "num": temp_faltas[u_max]}
    
    if temp_retrasos:
        u_max = max(temp_retrasos, key=temp_retrasos.get)
        max_retrasos = {"nombre": u_max, "num": temp_retrasos[u_max]}

    cur.execute("SELECT id, nombre FROM alumnos WHERE grupo_id = ? ORDER BY nombre", (grupo_id,))
    alumnos = cur.fetchall()

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()

    add_header(elements, styles, f"Informe Global del Grupo - Trimestre {trimestre}", colegio, fecha_inf)
    if tutor_nombre:
        elements.append(Paragraph(f"Tutor/a: {tutor_nombre}", styles['Heading2']))
    elements.append(Spacer(1, 20))

    if row_grupo:
        if row_grupo["observaciones"]:
            txt_formatted = row_grupo["observaciones"].replace('\n', '<br/>')
            elements.append(Paragraph("Valoración General:", styles['Heading3']))
            elements.append(Paragraph(txt_formatted, styles['BodyText']))
            elements.append(Spacer(1, 10))
        
        dificultades = ""
        try:
            dificultades = row_grupo["dificultades"] or ""
        except (IndexError, KeyError):
            pass
        if dificultades:
            txt_formatted = dificultades.replace('\n', '<br/>')
            elements.append(Paragraph("Dificultades Encontradas:", styles['Heading3']))
            elements.append(Paragraph(txt_formatted, styles['BodyText']))
            elements.append(Spacer(1, 10))
        
        if row_grupo["propuestas_mejora"]:
            txt_formatted = row_grupo["propuestas_mejora"].replace('\n', '<br/>')
            elements.append(Paragraph("Propuestas de Mejora:", styles['Heading3']))
            elements.append(Paragraph(txt_formatted, styles['BodyText']))
            elements.append(Spacer(1, 10))
        
        if row_grupo["conclusion"]:
            txt_formatted = row_grupo["conclusion"].replace('\n', '<br/>')
            elements.append(Paragraph("Conclusión:", styles['Heading3']))
            elements.append(Paragraph(txt_formatted, styles['BodyText']))
            elements.append(Spacer(1, 10))
    else:
        elements.append(Paragraph("No hay valoración grupal registrada.", styles['Normal']))
    elements.append(Spacer(1, 15))

    elements.append(Paragraph("Análisis de Promoción (Resultados):", styles['Heading3']))

    def calc_pct(n):
        return f"{round(n * 100 / total_alumnos, 1)}%" if total_alumnos > 0 else "0%"
    
    def calc_pct_inf(n, evaluados):
        return f"{round(n * 100 / evaluados, 1)}%" if evaluados > 0 else "0%"

    if es_infantil:
        periodo = f"T{trimestre}"
        # Obtener escala predominante para etiquetas
        cur.execute("""
            SELECT a.tipo_escala 
            FROM grupos g
            JOIN areas a ON g.etapa_id = a.etapa_id
            WHERE g.id = ? AND a.tipo_escala LIKE 'INFANTIL_%'
            LIMIT 1
        """, (grupo_id,))
        esc_row = cur.fetchone()
        scale_label = esc_row["tipo_escala"] if esc_row else "INFANTIL_NI_EP_C"
        
        is_pa_ma = scale_label == "INFANTIL_PA_A_MA"
        l1, l2, l3 = ("PA", "AD", "MA") if is_pa_ma else ("NI", "EP", "C")
        t1, t2, t3 = ("Parcialmente Adquirido", "Adquirido", "Muy Adquirido") if is_pa_ma else ("No Iniciado", "En proceso", "Conseguido")

        # CORRECCIÓN CRÍTICA: Query con parámetros correctos
        if area_id:
            # Con filtro de área: 6 marcadores, 6 parámetros
            query_inf = """
                SELECT e.alumno_id, e.nota
                FROM evaluaciones e
                JOIN alumnos a ON e.alumno_id = a.id
                JOIN areas ar ON e.area_id = ar.id
                WHERE e.trimestre = ? AND a.grupo_id = ? AND ar.id = ?
                UNION ALL
                SELECT ec.alumno_id, ec.nota
                FROM evaluacion_criterios ec
                JOIN alumnos a ON ec.alumno_id = a.id
                JOIN criterios c ON ec.criterio_id = c.id
                JOIN areas ar ON c.area_id = ar.id
                WHERE ec.periodo = ? AND a.grupo_id = ? AND ar.id = ?
            """
            params_inf = [trimestre, grupo_id, area_id, periodo, grupo_id, area_id]
        else:
            # Sin filtro de área: 4 marcadores, 4 parámetros
            query_inf = """
                SELECT e.alumno_id, e.nota
                FROM evaluaciones e
                JOIN alumnos a ON e.alumno_id = a.id
                JOIN areas ar ON e.area_id = ar.id
                WHERE e.trimestre = ? AND a.grupo_id = ?
                UNION ALL
                SELECT ec.alumno_id, ec.nota
                FROM evaluacion_criterios ec
                JOIN alumnos a ON ec.alumno_id = a.id
                JOIN criterios c ON ec.criterio_id = c.id
                JOIN areas ar ON c.area_id = ar.id
                WHERE ec.periodo = ? AND a.grupo_id = ?
            """
            params_inf = [trimestre, grupo_id, periodo, grupo_id]
        
        cur.execute(query_inf, params_inf)
        notas_inf = cur.fetchall()
        
        evaluados_infantil = 0
        infantil_map = {l3: 0, l2: 0, l1: 0}
        alumnos_medias = {}
        
        for row in notas_inf:
            aid = row["alumno_id"]
            if aid not in alumnos_medias:
                alumnos_medias[aid] = []
            alumnos_medias[aid].append(row["nota"])
        
        for aid, notas in alumnos_medias.items():
            notas_validas = [n for n in notas if n is not None]
            media_alumno = sum(notas_validas) / len(notas_validas) if notas_validas else 0
            evaluados_infantil += 1
            if media_alumno >= 2.5:
                infantil_map[l3] += 1
            elif media_alumno >= 1.5:
                infantil_map[l2] += 1
            else:
                infantil_map[l1] += 1

        data_prom = [
            [f"{t3} ({l3})", f"{infantil_map[l3]} ({calc_pct_inf(infantil_map[l3], evaluados_infantil)})"],
            [f"{t2} ({l2})", f"{infantil_map[l2]} ({calc_pct_inf(infantil_map[l2], evaluados_infantil)})"],
            [f"{t1} ({l1})", f"{infantil_map[l1]} ({calc_pct_inf(infantil_map[l1], evaluados_infantil)})"],
            ["No Evaluados", f"{total_alumnos - evaluados_infantil}"]
        ]
    else:
        data_prom = [
            ["Todo Aprobado", f"{susp_map[0]} ({calc_pct(susp_map[0])})"],
            ["1 Suspenso", f"{susp_map[1]} ({calc_pct(susp_map[1])})"],
            ["2 Suspensos", f"{susp_map[2]} ({calc_pct(susp_map[2])})"],
            ["+2 Suspensos", f"{susp_map[3]} ({calc_pct(susp_map[3])})"]
        ]
    
    t_prom = Table(data_prom, colWidths=[4*cm, 4*cm])
    t_prom.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 1, colors.grey),
        ('BACKGROUND', (0,0), (0,-1), colors.lightgrey)
    ]))
    elements.append(t_prom)
    elements.append(Spacer(1, 15))

    elements.append(Paragraph("Datos de Asistencia Destacados:", styles['Heading3']))
    elements.append(Paragraph(f"• Alumno/a con más faltas: <b>{max_faltas['nombre']}</b> ({max_faltas['num']} faltas)", styles['Normal']))
    elements.append(Paragraph(f"• Alumno/a con más retrasos: <b>{max_retrasos['nombre']}</b> ({max_retrasos['num']} retrasos)", styles['Normal']))

    equipo_docente = ""
    if row_grupo:
        try:
            equipo_docente = row_grupo["equipo_docente"] if "equipo_docente" in row_grupo.keys() else ""
        except:
            pass
        
    if equipo_docente:
        elements.append(Spacer(1, 20))
        elements.append(Paragraph("Miembros del Equipo Docente", styles['Heading3']))
        
        firmantes = [line.strip() for line in equipo_docente.replace('\r', '').split('\n') if line.strip()]
        if firmantes:
            sig_data = [["Docente", "Firma"]]
            for f in firmantes:
                sig_data.append([Paragraph(f, styles['Normal']), ""])
                
            t_sig = Table(sig_data, colWidths=[8*cm, 8*cm], rowHeights=[1*cm] + [1.5*cm]*len(firmantes))
            t_sig.setStyle(TableStyle([
                ('GRID', (0,0), (-1,-1), 1, colors.black),
                ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('ALIGN', (0,0), (-1,0), 'CENTER'),
            ]))
            elements.append(t_sig)

    elements.append(PageBreak())

    for al in alumnos:
        query_crit = """
            SELECT ar.nombre as area, s.nombre as sda, c.codigo as criterio_codigo, c.descripcion as criterio_desc, e.nota, ar.tipo_escala, e.nivel, c.comentario_base as base
            FROM evaluaciones e
            JOIN areas ar ON e.area_id = ar.id
            JOIN sda s ON e.sda_id = s.id
            JOIN criterios c ON e.criterio_id = c.id
            WHERE e.alumno_id = ? AND e.trimestre = ?
        """
        params_crit = [al["id"], trimestre]
        if area_id:
            query_crit += " AND ar.id = ?"
            params_crit.append(area_id)
            
        query_crit += """
            UNION ALL
            SELECT ar.nombre as area, 'Criterios Directos' as sda, c.codigo as criterio_codigo, c.descripcion as criterio_desc, ec.nota, ar.tipo_escala, ec.nivel, c.comentario_base as base
            FROM evaluacion_criterios ec
            JOIN criterios c ON ec.criterio_id = c.id
            JOIN areas ar ON c.area_id = ar.id
            WHERE ec.alumno_id = ? AND ec.periodo = ?
        """
        params_crit.extend([al["id"], periodo])
        if area_id:
            query_crit += " AND ar.id = ?"
            params_crit.append(area_id)
            
        query_crit += " ORDER BY area, sda, criterio_codigo"
        
        cur.execute(query_crit, params_crit)
        notas_criterios = cur.fetchall()
        
        add_header(elements, styles, f"Informe Trimestral - Trimestre {trimestre}", colegio, fecha_inf)
        elements.append(Paragraph(f"Alumno: {al['nombre']}", styles['Heading2']))
        if tutor_nombre:
            elements.append(Paragraph(f"Tutor/a: {tutor_nombre}", styles['Normal']))
        elements.append(Spacer(1, 12))
        
        cur.execute("""
            SELECT estado, COUNT(*) 
            FROM asistencia 
            WHERE alumno_id = ?
            GROUP BY estado
        """, (al['id'],))
        asist_data = dict(cur.fetchall())

        cur.execute("""
            SELECT a.id, a.nombre, ROUND(AVG(val.nota), 2) as media, a.tipo_escala
            FROM (
                SELECT area_id, nota FROM evaluaciones WHERE alumno_id = ? AND trimestre = ?
                UNION ALL
                SELECT c.area_id, ec.nota 
                FROM evaluacion_criterios ec
                JOIN criterios c ON ec.criterio_id = c.id
                WHERE ec.alumno_id = ? AND ec.periodo = ?
            ) val
            JOIN areas a ON val.area_id = a.id
            GROUP BY a.id, a.nombre, a.tipo_escala
        """, (al['id'], trimestre, al['id'], periodo))
        notas_area = cur.fetchall()

        cur.execute("""
            SELECT fecha, estado
            FROM asistencia
            WHERE alumno_id = ? AND estado IN ('retraso', 'falta_justificada', 'falta_no_justificada')
            AND fecha BETWEEN ? AND ?
            ORDER BY fecha
        """, (al['id'], start_date, end_date))
        faltas_detalle = cur.fetchall()
        
        cur.execute("SELECT texto FROM informe_individual WHERE alumno_id = ? AND trimestre = ?", (al['id'], trimestre))
        obs_row = cur.fetchone()
        obs_text = obs_row["texto"] if obs_row else ""
        
        elements.append(Paragraph("Resumen de Asistencia", styles['Heading3']))
        data_asist = [
            ["Estado", "Días"],
            [Paragraph('<font color="#ffc107">●</font> Retraso', styles['Normal']), asist_data.get('retraso', 0)],
            [Paragraph('<font color="#17a2b8">●</font> Falta Justificada', styles['Normal']), asist_data.get('falta_justificada', 0)],
            [Paragraph('<font color="#dc3545">●</font> Falta No Justificada', styles['Normal']), asist_data.get('falta_no_justificada', 0)],
        ]
        t_asist = Table(data_asist, colWidths=[6*cm, 2*cm])
        t_asist.setStyle(TableStyle([
            ('GRID', (0,0), (-1,-1), 1, colors.black),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE')
        ]))
        elements.append(t_asist)
        elements.append(Spacer(1, 12))

        if faltas_detalle:
            fechas_str = []
            for row in faltas_detalle:
                f = row["fecha"]
                e = row["estado"]
                try:
                    d_obj = datetime.strptime(f, '%Y-%m-%d')
                    d_str = d_obj.strftime('%d/%m')
                except:
                    d_str = f

                if e == 'retraso':
                    color, tipo = "#ffc107", "R"
                elif e == 'falta_justificada':
                    color, tipo = "#17a2b8", "F"
                else:
                    color, tipo = "#dc3545", "F"

                fechas_str.append(f'<font color="{color}">{d_str}({tipo})</font>')
            elements.append(Paragraph("<b>Detalle (Día/Tipo):</b> " + ", ".join(fechas_str), styles['Normal']))
            elements.append(Spacer(1, 12))

        elements.append(Paragraph("Rendimiento Académico", styles['Heading3']))
        
        area_sda_map = {}
        for area, sda, crit_cod, crit_desc, nota, escala, nivel, base in notas_criterios:
            if area not in area_sda_map:
                area_sda_map[area] = {}
            if sda not in area_sda_map[area]:
                area_sda_map[area][sda] = []
            area_sda_map[area][sda].append((f"{crit_cod}", crit_desc or "", nota, escala))
        
        data_notas = [["Área", "Nota Media", "Detalle SDA y Criterios"]]
        for aid, area_nombre, nota_media, tipo_escala in notas_area:
            sda_dict = area_sda_map.get(area_nombre, {})
            detail_lines = []
            for sda_name, criterios in sda_dict.items():
                eval_notas = [c[2] for c in criterios if c[2] is not None]
                sda_avg = round(sum(eval_notas) / len(eval_notas), 2) if eval_notas else 0
                fmt_sda_avg = format_nota(sda_avg, tipo_escala)
                detail_lines.append(f"<b>{sda_name}</b>: {fmt_sda_avg}")
                for crit_cod, crit_desc, nota, escala in criterios:
                    desc_txt = f" – {crit_desc}" if crit_desc else ""
                    detail_lines.append(f"  • {crit_cod}{desc_txt}: {format_nota(nota, escala)}")
            detail_str = "<br/>".join(detail_lines) if detail_lines else "Sin datos"
            data_notas.append([
                Paragraph(area_nombre, styles['Normal']), 
                str(format_nota(nota_media, tipo_escala)), 
                Paragraph(detail_str, styles['Normal'])
            ])
            
        t = Table(data_notas, colWidths=[4*cm, 2.5*cm, 10*cm])
        t.setStyle(TableStyle([
            ('GRID', (0,0), (-1,-1), 1, colors.black),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('BACKGROUND', (0,0), (-1,0), colors.lightgrey)
        ]))
        elements.append(t)
        
        from reportlab.platypus import Image as RLImage
        chart_buf = generar_grafica_rendimiento(notas_area, al['nombre'], es_infantil)
        if chart_buf:
            elements.append(Spacer(1, 8))
            elements.append(RLImage(chart_buf, width=16*cm, height=max(4, len(notas_area) * 0.7)*cm))
        
        if obs_text:
            elements.append(Spacer(1, 12))
            elements.append(Paragraph("<b>Observaciones:</b>", styles['Heading3']))
            elements.append(Paragraph(obs_text, styles['Normal']))

        elements.append(Spacer(1, 12))
        elements.append(Paragraph("Observaciones Pedagógicas", styles['Heading3']))
        comentarios_por_area = {}
        for area, sda, crit_cod, crit_desc, nota, escala, nivel, base in notas_criterios:
            if nivel:
                comment = ""
                if (escala and escala.startswith("INFANTIL_")):
                    is_pama = (escala == "INFANTIL_PA_A_MA")
                    if nivel == 1:
                        comment = f"• <b>{crit_cod}</b>: En proceso inicial de adquirir {crit_desc}." if is_pama else f"• <b>{crit_cod}</b>: No ha iniciado {crit_desc}."
                    elif nivel == 2:
                        comment = f"• <b>{crit_cod}</b>: Ha progresado adecuadamente en {crit_desc}." if is_pama else f"• <b>{crit_cod}</b>: Está en proceso de adquirir {crit_desc}."
                    elif nivel >= 3:
                        comment = f"• <b>{crit_cod}</b>: Ha adquirido plenamente {crit_desc}." if is_pama else f"• <b>{crit_cod}</b>: Ha conseguido {crit_desc}."
                else:
                    if nivel == 1:
                        comment = f"• <b>{crit_cod}</b>: Necesita apoyo en {crit_desc}."
                    elif nivel == 2:
                        comment = f"• <b>{crit_cod}</b>: Está en proceso de mejorar en {crit_desc}."
                    elif nivel == 3:
                        comment = f"• <b>{crit_cod}</b>: Comprende y aplica adecuadamente {crit_desc}."
                    elif nivel >= 4:
                        comment = f"• <b>{crit_cod}</b>: Destaca especialmente en {crit_desc}."
                
                if base:
                    comment += f" {base}"
                    
                if area not in comentarios_por_area:
                    comentarios_por_area[area] = []
                comentarios_por_area[area].append(comment)
            
        if comentarios_por_area:
            for area, comentarios in comentarios_por_area.items():
                elements.append(Paragraph(f"<b><u>{area}</u></b>", styles['Normal']))
                for c in comentarios:
                    elements.append(Paragraph(c, styles['Normal']))
                elements.append(Spacer(1, 4))
        else:
            elements.append(Paragraph("Sin observaciones registradas.", styles['Normal']))

        elements.append(Spacer(1, 40))
        if tutor_nombre:
            firma_data = [
                [Paragraph(f"Fdo: {tutor_nombre}", styles['Normal'])],
                [Paragraph(f"({cargo_tutor})", styles['Normal'])]
            ]
            t_firma = Table(firma_data, colWidths=[10*cm])
            t_firma.setStyle(TableStyle([
                ('LEFTPADDING', (0,0), (-1,-1), 0),
                ('TOPPADDING', (0,0), (-1,-1), 0),
                ('BOTTOMPADDING', (0,0), (-1,-1), 2),
            ]))
            elements.append(t_firma)

        elements.append(PageBreak())
    
    doc.build(elements)
    buffer.seek(0)

    filename = f"Informe_General_{grupo_nombre.replace(' ', '_')}_T{trimestre}.pdf"

    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/pdf'
    )

# ==============================================================================
# OTRAS RUTAS (Mantenidas igual - sin cambios críticos)
# ==============================================================================

@informes_bp.route("/api/informe/preview_diario/<int:alumno_id>")
def informe_preview_diario(alumno_id):
    conn = get_db()
    cur = conn.cursor()
    
    cur.execute("SELECT nombre FROM alumnos WHERE id = ?", (alumno_id,))
    row_alumno = cur.fetchone()
    if not row_alumno:
        return jsonify({"ok": False, "error": "Alumno no encontrado"}), 404
    
    nombre_alumno = row_alumno["nombre"]

    cur.execute("""
        SELECT o.id, o.fecha, o.texto, o.foto, ar.nombre as area_nombre
        FROM observaciones o
        LEFT JOIN areas ar ON o.area_id = ar.id
        WHERE o.alumno_id = ?
        ORDER BY o.fecha DESC, o.id DESC
    """, (alumno_id,))
    rows = cur.fetchall()

    grouped = {}
    for r in rows:
        fecha = r["fecha"]
        if fecha not in grouped:
            grouped[fecha] = []
        grouped[fecha].append({
            "id": r["id"],
            "texto": r["texto"],
            "foto": r["foto"] or "",
            "area": r["area_nombre"] or "General"
        })
    
    data = []
    for fecha in sorted(grouped.keys(), reverse=True):
        y, m, d = fecha.split('-')
        fecha_fmt = f"{d}/{m}/{y}"
        data.append({
            "fecha": fecha_fmt,
            "raw_fecha": fecha,
            "observaciones": grouped[fecha]
        })
    
    return jsonify({
        "ok": True,
        "nombre_alumno": nombre_alumno,
        "data": data
    })

@informes_bp.route("/api/informe/pdf_diario/<int:alumno_id>")
def informe_pdf_diario(alumno_id):
    conn = get_db()
    cur = conn.cursor()
    
    cur.execute("SELECT nombre FROM alumnos WHERE id = ?", (alumno_id,))
    row_al = cur.fetchone()
    if not row_al:
        return "Alumno no encontrado", 404
    alumno = row_al["nombre"]

    cur.execute("""
        SELECT o.fecha, ar.nombre, o.texto
        FROM observaciones o
        LEFT JOIN areas ar ON o.area_id = ar.id
        WHERE o.alumno_id = ?
        ORDER BY o.fecha DESC
    """, (alumno_id,))
    obs = cur.fetchall()

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()

    add_header(elements, styles, f"Diario de Clase: {alumno}")
    elements.append(Spacer(1, 12))

    tag_colors = {
        '#SinMaterial': '#a33636',
        '#FaltaTarea': '#d47500',
        '#BuenaActitud': '#28a745',
        '#MalComportamiento': '#dc3545',
        '#Participa': '#17a2b8'
    }

    def format_tags(text):
        if not text:
            return ""
        def repl(m):
            t = m.group(0)
            c = tag_colors.get(t, '#007bff')
            return f'<font color="{c}"><b>{t}</b></font>'
        return re.sub(r'#\w+', repl, text)

    for fecha, area, texto in obs:
        area_str = f" ({area})" if area else ""
        formatted_text = format_tags(texto)
        elements.append(Paragraph(f"<b>{fecha}{area_str}:</b> {formatted_text}", styles['BodyText']))
        elements.append(Spacer(1, 6))
    
    doc.build(elements)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"Diario_{alumno}.pdf",
        mimetype='application/pdf'
    )

@informes_bp.route("/api/informe/grupo_obs", methods=["GET", "POST"])
def grupo_obs():
    trimestre = request.args.get("trimestre") if request.method == "GET" else request.json.get("trimestre")
    if not trimestre:
        return jsonify({"ok": False, "error": "Falta trimestre"}), 400
    
    conn = get_db()
    cur = conn.cursor()

    if request.method == "POST":
        d = request.json
        obs = d.get("observaciones", "")
        prop = d.get("propuestas_mejora", "")
        conc = d.get("conclusion", "")
        eq_doc = d.get("equipo_docente", "")
        dif = d.get("dificultades", "")
        
        grupo_id = session.get('active_group_id')
        cur.execute("""
            INSERT INTO informe_grupo (grupo_id, trimestre, observaciones, propuestas_mejora, conclusion, equipo_docente, dificultades)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(grupo_id, trimestre) DO UPDATE SET
                observaciones = excluded.observaciones,
                propuestas_mejora = excluded.propuestas_mejora,
                conclusion = excluded.conclusion,
                equipo_docente = excluded.equipo_docente,
                dificultades = excluded.dificultades
        """, (grupo_id, trimestre, obs, prop, conc, eq_doc, dif))
        conn.commit()
        return jsonify({"ok": True})
    else:
        grupo_id = session.get('active_group_id')
        
        cur.execute("SELECT equipo_docente FROM grupos WHERE id = ?", (grupo_id,))
        g_row = cur.fetchone()
        eq_doc_grupo = ""
        if g_row:
            try:
                eq_doc_grupo = g_row["equipo_docente"] if "equipo_docente" in g_row.keys() and g_row["equipo_docente"] else ""
            except:
                eq_doc_grupo = g_row[0] if (type(g_row) is tuple) or (g_row and g_row[0]) else ""
        
        cur.execute("SELECT * FROM informe_grupo WHERE trimestre = ? AND grupo_id = ?", (trimestre, grupo_id))
        row = cur.fetchone()
        
        if row:
            equipo = ""
            try:
                equipo = row["equipo_docente"] if "equipo_docente" in row.keys() else ""
            except:
                pass
            res = {
                "observaciones": row["observaciones"],
                "propuestas_mejora": row["propuestas_mejora"],
                "conclusion": row["conclusion"],
                "equipo_docente": equipo,
                "equipo_docente_grupo": eq_doc_grupo
            }
            try:
                if "dificultades" in row.keys():
                    res["dificultades"] = row["dificultades"]
                else:
                    res["dificultades"] = ""
            except:
                try:
                    res["dificultades"] = row[7] if len(row) > 7 else ""
                except:
                    res["dificultades"] = ""
            return jsonify(res)
        return jsonify({"equipo_docente_grupo": eq_doc_grupo, "dificultades": ""})

@informes_bp.route("/api/informe/grupo_obs_delete", methods=["POST"])
def grupo_obs_delete():
    data = request.json or {}
    trimestre = data.get("trimestre")
    if not trimestre:
        return jsonify({"ok": False, "error": "Falta trimestre"}), 400
    
    grupo_id = session.get('active_group_id')
    if not grupo_id:
        return jsonify({"ok": False, "error": "No hay grupo activo"}), 400

    conn = get_db()
    cur = conn.cursor()
    try:
        cur.execute("DELETE FROM informe_grupo WHERE grupo_id = ? AND trimestre = ?", (grupo_id, trimestre))
        conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        conn.rollback()
        return jsonify({"ok": False, "error": str(e)}), 500

@informes_bp.route("/api/informe/historial_status")
def historial_status():
    grupo_id = session.get('active_group_id')
    if not grupo_id:
        return jsonify([])
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT trimestre FROM informe_grupo WHERE grupo_id = ?", (grupo_id,))
    rows = cur.fetchall()
    exists = [r["trimestre"] for r in rows]
    
    res = []
    for t in [1, 2, 3]:
        res.append({
            "trimestre": t,
            "exists": t in exists
        })
    return jsonify(res)

@informes_bp.route("/api/informe/grupo_data")
def grupo_data():
    trimestre = request.args.get("trimestre", "1")
    conn = get_db()
    cur = conn.cursor()
    grupo_id = session.get('active_group_id')

    cur.execute("SELECT COUNT(*) FROM alumnos WHERE grupo_id = ?", (grupo_id,))
    total_alumnos = cur.fetchone()[0]

    cur.execute("""
        SELECT a.tipo_escala 
        FROM grupos g
        JOIN areas a ON g.etapa_id = a.etapa_id
        WHERE g.id = ? AND a.tipo_escala LIKE 'INFANTIL_%'
        LIMIT 1
    """, (grupo_id,))
    es_infantil = cur.fetchone() is not None

    media_general = 0
    total_evals = 0
    suspensos_map = {}

    if es_infantil:
        periodo = f"T{trimestre}"
        cur.execute("""
            SELECT ec.alumno_id, ec.nota
            FROM evaluacion_criterios ec
            JOIN alumnos a ON ec.alumno_id = a.id
            WHERE ec.periodo = ? AND a.grupo_id = ?
        """, (periodo, grupo_id))
        
        notas_inf = cur.fetchall()
        total_evals = len(notas_inf)
        if total_evals > 0:
            notas_validas = [row["nota"] for row in notas_inf if row["nota"] is not None]
            media_general = round(sum(notas_validas) / len(notas_validas), 2) if notas_validas else 0
            
            alumnos_medias = {}
            for row in notas_inf:
                aid = row["alumno_id"]
                if aid not in alumnos_medias:
                    alumnos_medias[aid] = []
                alumnos_medias[aid].append(row["nota"])
                
            for aid, notas in alumnos_medias.items():
                notas_validas = [n for n in notas if n is not None]
                media_alumno = sum(notas_validas) / len(notas_validas) if notas_validas else 0
                
                # Obtener escala de este alumno (simplificado a la del grupo)
                cur.execute("""
                    SELECT a.tipo_escala FROM alumnos al 
                    JOIN grupos g ON al.grupo_id = g.id
                    JOIN areas a ON g.etapa_id = a.etapa_id
                    WHERE al.id = ? AND a.tipo_escala LIKE 'INFANTIL_%' LIMIT 1
                """, (aid,))
                row_esc = cur.fetchone()
                esc = row_esc["tipo_escala"] if row_esc else "INFANTIL_NI_EP_C"
                l1, l2, l3 = ("PA", "AD", "MA") if esc == "INFANTIL_PA_A_MA" else ("NI", "EP", "C")

                if media_alumno >= 2.5:
                    suspensos_map[aid] = l3
                elif media_alumno >= 1.5:
                    suspensos_map[aid] = l2
                else:
                    suspensos_map[aid] = l1
    else:
        cur.execute("""
            SELECT COUNT(*), AVG(e.nota) 
            FROM evaluaciones e 
            JOIN alumnos a ON e.alumno_id = a.id 
            WHERE e.trimestre = ? AND a.grupo_id = ?
        """, (trimestre, grupo_id))
        row_evals = cur.fetchone()
        total_evals = row_evals[0]
        media_general = round(row_evals[1] or 0, 2)
        
        cur.execute("""
            SELECT alumno_id, COUNT(area_id) as num_suspensos
            FROM (
                SELECT alumno_id, area_id, AVG(nota) as media_area
                FROM evaluaciones
                WHERE trimestre = ?
                GROUP BY alumno_id, area_id
                HAVING media_area < 5
            ) val
            JOIN alumnos a ON val.alumno_id = a.id
            WHERE a.grupo_id = ?
            GROUP BY alumno_id
        """, (trimestre, grupo_id))
        suspensos_map = {row['alumno_id']: row['num_suspensos'] for row in cur.fetchall()}
    
    year = date.today().year
    if trimestre == "1":
        start_date, end_date = f"{year-1}-09-01", f"{year-1}-12-31"
    elif trimestre == "2":
        start_date, end_date = f"{year}-01-01", f"{year}-03-31"
    else:
        start_date, end_date = f"{year}-04-01", f"{year}-06-30"
    
    cur.execute("""
        SELECT asi.estado, COUNT(*)
        FROM asistencia asi
        JOIN alumnos a ON a.id = asi.alumno_id
        WHERE asi.fecha BETWEEN ? AND ? AND a.grupo_id = ?
        GROUP BY asi.estado
    """, (start_date, end_date, grupo_id))

    asist_stats = dict(cur.fetchall())

    cur.execute("SELECT id FROM alumnos WHERE grupo_id = ?", (grupo_id,))
    all_ids = [r[0] for r in cur.fetchall()]

    distribucion = {0: 0, 1: 0, 2: 0, 3: 0}
    infantil_dist = {'C': 0, 'EP': 0, 'NI': 0}
    evaluados_count = 0

    if es_infantil:
        for aid in all_ids:
            cat = suspensos_map.get(aid)
            if cat in infantil_dist:
                infantil_dist[cat] += 1
                evaluados_count += 1
    else:
        for aid in all_ids:
            num = suspensos_map.get(aid, 0)
            if num > 2:
                distribucion[3] += 1
            else:
                distribucion[num] += 1
        
    def pct(n):
        return round(n * 100.0 / total_alumnos, 1) if total_alumnos > 0 else 0
    
    def pct_inf(n):
        return round(n * 100.0 / evaluados_count, 1) if evaluados_count > 0 else 0

    promocion_data = {}
    if es_infantil:
        # Etiquetas dinámicas para el JSON de respuesta
        cur.execute("""
            SELECT a.tipo_escala FROM grupos g
            JOIN areas a ON g.etapa_id = a.etapa_id
            WHERE g.id = ? AND a.tipo_escala LIKE 'INFANTIL_%' LIMIT 1
        """, (grupo_id,))
        esc_row = cur.fetchone()
        esc = esc_row["tipo_escala"] if esc_row else "INFANTIL_NI_EP_C"
        l1, l2, l3 = ("PA", "AD", "MA") if esc == "INFANTIL_PA_A_MA" else ("NI", "EP", "C")

        promocion_data = {
            "is_infantil": True,
            "labels": {"L1": l1, "L2": l2, "L3": l3},
            "L3": {"num": infantil_dist.get(l3, 0), "pct": pct_inf(infantil_dist.get(l3, 0))},
            "L2": {"num": infantil_dist.get(l2, 0), "pct": pct_inf(infantil_dist.get(l2, 0))},
            "L1": {"num": infantil_dist.get(l1, 0), "pct": pct_inf(infantil_dist.get(l1, 0))},
            "NE": {"num": total_alumnos - evaluados_count, "pct": 0}
        }
    else:
        promocion_data = {
            "is_infantil": False,
            "todo": {"num": distribucion[0], "pct": pct(distribucion[0])},
            "una": {"num": distribucion[1], "pct": pct(distribucion[1])},
            "dos": {"num": distribucion[2], "pct": pct(distribucion[2])},
            "mas_de_dos": {"num": distribucion[3], "pct": pct(distribucion[3])}
        }

    return jsonify({
        "generales": {
            "total_alumnos": total_alumnos,
            "media_general": media_general,
            "total_evals": total_evals,
            "is_infantil": es_infantil
        },
        "asistencia": {
            "total_faltas": asist_stats.get('falta_justificada', 0) + asist_stats.get('falta_no_justificada', 0),
            "f_justificada": asist_stats.get('falta_justificada', 0),
            "f_no_justificada": asist_stats.get('falta_no_justificada', 0),
            "total_retrasos": asist_stats.get('retraso', 0)
        },
        "promocion": promocion_data
    })

@informes_bp.route("/api/informe/asistencia_detalle")
def asistencia_detalle():
    trimestre = request.args.get("trimestre", "1")
    estado = request.args.get("estado")
    
    year = date.today().year
    if trimestre == "1":
        start_date, end_date = f"{year-1}-09-01", f"{year-1}-12-31"
    elif trimestre == "2":
        start_date, end_date = f"{year}-01-01", f"{year}-03-31"
    else:
        start_date, end_date = f"{year}-04-01", f"{year}-06-30"

    grupo_id = session.get('active_group_id')
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT a.nombre, asi.fecha
        FROM asistencia asi
        JOIN alumnos a ON asi.alumno_id = a.id
        WHERE asi.estado = ? AND asi.fecha BETWEEN ? AND ? AND a.grupo_id = ?
        ORDER BY asi.fecha DESC
    """, (estado, start_date, end_date, grupo_id))
    rows = cur.fetchall()

    return jsonify([{"nombre": r["nombre"], "fecha": r["fecha"]} for r in rows])

@informes_bp.route("/api/informe/observacion", methods=["GET", "POST"])
def observacion_individual():
    if request.method == "POST":
        d = request.json
        alumno_id = d.get("alumno_id")
        trimestre = d.get("trimestre")
        texto = d.get("texto")
        
        conn = get_db()
        cur = conn.cursor()
        
        cur.execute("""
            INSERT INTO informe_individual (alumno_id, trimestre, texto)
            VALUES (?, ?, ?)
            ON CONFLICT(alumno_id, trimestre) DO UPDATE SET
                texto = excluded.texto,
                fecha_actualizacion = CURRENT_TIMESTAMP
        """, (alumno_id, trimestre, texto))
        
        conn.commit()
        return jsonify({"ok": True})
    else:
        alumno_id = request.args.get("alumno_id")
        trimestre = request.args.get("trimestre")
        
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT texto FROM informe_individual WHERE alumno_id = ? AND trimestre = ?", (alumno_id, trimestre))
        row = cur.fetchone()
        
        return jsonify({"texto": row["texto"] if row else ""})

@informes_bp.route("/api/informe/asistencia_alumno")
def asistencia_alumno_stats():
    alumno_id = request.args.get("alumno_id")
    trimestre = request.args.get("trimestre")
    
    year = date.today().year
    if trimestre == "1":
        start_date, end_date = f"{year-1}-09-01", f"{year-1}-12-31"
    elif trimestre == "2":
        start_date, end_date = f"{year}-01-01", f"{year}-03-31"
    else:
        start_date, end_date = f"{year}-04-01", f"{year}-06-30"

    conn = get_db()
    cur = conn.cursor()

    cur.execute("""
        SELECT estado, COUNT(*)
        FROM asistencia
        WHERE alumno_id = ? AND fecha BETWEEN ? AND ?
        GROUP BY estado
    """, (alumno_id, start_date, end_date))

    stats_rows = cur.fetchall()
    stats = {row["estado"]: row[1] for row in stats_rows}

    cur.execute("""
        SELECT strftime('%m', fecha) as mes, estado, COUNT(*) as cuenta
        FROM asistencia
        WHERE alumno_id = ? AND fecha BETWEEN ? AND ?
        AND estado IN ('retraso', 'falta_justificada', 'falta_no_justificada')
        GROUP BY mes, estado
        ORDER BY mes
    """, (alumno_id, start_date, end_date))

    mensual_rows = cur.fetchall()
    mensual = {}
    nombres_meses = {
        "01": "Enero", "02": "Febrero", "03": "Marzo", "04": "Abril",
        "05": "Mayo", "06": "Junio", "07": "Julio", "08": "Agosto",
        "09": "Septiembre", "10": "Octubre", "11": "Noviembre", "12": "Diciembre"
    }

    for row in mensual_rows:
        m = nombres_meses.get(row["mes"], row["mes"])
        if m not in mensual:
            mensual[m] = {"retrasos": 0, "faltas": 0}
        
        if row["estado"] == "retraso":
            mensual[m]["retrasos"] += row["cuenta"]
        else:
            mensual[m]["faltas"] += row["cuenta"]
        
    return jsonify({
        "retrasos": stats.get("retraso", 0),
        "f_justificada": stats.get("falta_justificada", 0),
        "f_no_justificada": stats.get("falta_no_justificada", 0),
        "total_faltas": stats.get("falta_justificada", 0) + stats.get("falta_no_justificada", 0),
        "mensual": mensual
    })

@informes_bp.route("/api/informe/excel_individual")
def excel_individual():
    alumno_id = request.args.get("alumno_id")
    trimestre = request.args.get("trimestre")
    if not alumno_id or not trimestre:
        return "Faltan parámetros", 400

    conn = get_db()
    cur = conn.cursor()

    cur.execute("SELECT nombre FROM alumnos WHERE id = ?", (alumno_id,))
    row_al = cur.fetchone()
    if not row_al:
        return "Alumno no encontrado", 404
    alumno_nombre = row_al["nombre"]

    periodo = f"T{trimestre}"

    cur.execute("""
        SELECT a.nombre as area, s.nombre as sda, c.codigo, c.descripcion, e.nota, a.tipo_escala
        FROM evaluaciones e
        JOIN sda s ON e.sda_id = s.id
        JOIN areas a ON e.area_id = a.id
        JOIN criterios c ON e.criterio_id = c.id
        WHERE e.alumno_id = ? AND e.trimestre = ?
        ORDER BY a.nombre, s.nombre, CAST(SUBSTR(c.codigo, INSTR(c.codigo, '.')+1) AS INTEGER), c.codigo
    """, (alumno_id, trimestre))
    rows_sda = cur.fetchall()

    cur.execute("""
        SELECT a.nombre as area, 'Evaluación Directa' as sda, c.codigo, c.descripcion, ec.nota, a.tipo_escala
        FROM evaluacion_criterios ec
        JOIN criterios c ON ec.criterio_id = c.id
        JOIN areas a ON c.area_id = a.id
        WHERE ec.alumno_id = ? AND ec.periodo = ?
        ORDER BY a.nombre, CAST(SUBSTR(c.codigo, INSTR(c.codigo, '.')+1) AS INTEGER), c.codigo
    """, (alumno_id, periodo))
    rows_dir = cur.fetchall()

    all_rows = list(rows_sda) + list(rows_dir)

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return "openpyxl no instalado", 500

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"T{trimestre} - {alumno_nombre[:20]}"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F3864")
    area_fill = PatternFill("solid", fgColor="D5E8F0")
    area_font = Font(bold=True)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A1:F1")
    ws["A1"] = f"Notas - {alumno_nombre} - {trimestre}º Trimestre"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = center

    headers = ["Área", "SDA / Modalidad", "Código Criterio", "Descripción", "Nota (1-10)", "Escala"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    current_area = None
    row_num = 3
    for r in all_rows:
        area = r["area"]
        if area != current_area:
            ws.merge_cells(f"A{row_num}:F{row_num}")
            cell = ws.cell(row=row_num, column=1, value=f"📚 {area}")
            cell.font = area_font
            cell.fill = area_fill
            cell.border = border
            row_num += 1
            current_area = area

        nota = r["nota"]
        escala = r["tipo_escala"]

        if (escala and escala.startswith("INFANTIL_")):
            n = round(nota) if nota <= 3 else (1 if nota <= 3.5 else (2 if nota <= 6.5 else 3))
            nota_display = {1: 1, 2: 5, 3: 10}.get(n, nota)
            if escala == "INFANTIL_PA_A_MA":
                escala_txt = {1: "PA", 2: "AD", 3: "MA"}.get(n, "—")
            else:
                escala_txt = {1: "NI", 2: "EP", 3: "C"}.get(n, "—")
        else:
            nota_display = round(nota, 2) if nota is not None else None
            escala_txt = "1-10"

        ws.cell(row=row_num, column=1, value="").border = border
        ws.cell(row=row_num, column=2, value=r["sda"]).border = border
        ws.cell(row=row_num, column=3, value=r["codigo"]).border = border
        ws.cell(row=row_num, column=4, value=r["descripcion"]).border = border
        nota_cell = ws.cell(row=row_num, column=5, value=nota_display)
        nota_cell.border = border
        nota_cell.alignment = center
        if nota_display is not None:
            if (escala and escala.startswith("INFANTIL_")):
                nota_cell.fill = PatternFill("solid", fgColor=("FADADD" if n == 1 else "FFF3CD" if n == 2 else "D5F5E3"))
            else:
                nota_cell.fill = PatternFill("solid", fgColor=("FADADD" if nota < 5 else "FFF3CD" if nota < 7 else "D5F5E3"))
        ws.cell(row=row_num, column=6, value=escala_txt).border = border
        row_num += 1

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 45
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"Notas_{alumno_nombre.replace(' ', '_')}_T{trimestre}.xlsx"
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@informes_bp.route("/api/informe/acta_oficial")
def acta_oficial():
    """Genera el Acta Oficial de Evaluación en PDF."""
    trimestre = request.args.get("trimestre", "2")
    tutor_nombre = request.args.get("tutor", "")
    fecha_str = request.args.get("fecha", "")
    hora_str = request.args.get("hora", "")
    grupo_id = session.get('active_group_id')
    if not grupo_id:
        return "No hay grupo activo", 400

    conn = get_db()
    cur = conn.cursor()

    # Obtener nombre del tutor desde configuración si no se pasa
    if not tutor_nombre:
        cur.execute("SELECT valor FROM config WHERE clave = 'nombre_tutor'")
        row = cur.fetchone()
        if row:
            tutor_nombre = row["valor"]

    cur.execute("SELECT g.nombre, g.curso, g.equipo_docente FROM grupos g WHERE g.id = ?", (grupo_id,))
    grupo = cur.fetchone()
    if not grupo:
        return "Grupo no encontrado", 404
    grupo_nombre = grupo["nombre"]
    equipo_docente_raw = grupo["equipo_docente"] or ""

    cur.execute("SELECT * FROM informe_grupo WHERE grupo_id = ? AND trimestre = ?", (grupo_id, trimestre))
    informe = cur.fetchone()
    valoracion = informe["observaciones"] if informe and informe["observaciones"] else ""
    dificultades = informe["dificultades"] if informe and "dificultades" in informe.keys() and informe["dificultades"] else ""
    propuestas = informe["propuestas_mejora"] if informe and informe["propuestas_mejora"] else ""
    otras_obs = informe["conclusion"] if informe and informe["conclusion"] else ""
    equipo_informe = informe["equipo_docente"] if informe and informe["equipo_docente"] else equipo_docente_raw

    cur.execute("SELECT COUNT(*) FROM alumnos WHERE grupo_id = ? AND deleted_at IS NULL", (grupo_id,))
    total_alumnos = cur.fetchone()[0]

    cur.execute("SELECT tipo_evaluacion FROM grupos WHERE id = ?", (grupo_id,))
    tipo = cur.fetchone()
    es_infantil = tipo and tipo["tipo_evaluacion"] == "infantil"

    periodo = f"T{trimestre}"
    aprobados = 0
    if es_infantil:
        cur.execute("""
            SELECT ec.alumno_id, AVG(ec.nota) as media
            FROM evaluacion_criterios ec
            JOIN alumnos a ON ec.alumno_id = a.id
            WHERE ec.periodo = ? AND a.grupo_id = ? AND a.deleted_at IS NULL
            GROUP BY ec.alumno_id
            HAVING media >= 2.5
        """, (periodo, grupo_id))
        aprobados = len(cur.fetchall())
    else:
        cur.execute("""
            SELECT alumno_id FROM (
                SELECT e.alumno_id, AVG(e.nota) as media_area
                FROM evaluaciones e
                JOIN alumnos a ON e.alumno_id = a.id
                WHERE e.trimestre = ? AND a.grupo_id = ? AND a.deleted_at IS NULL
                GROUP BY e.alumno_id, e.area_id
            ) sub GROUP BY alumno_id HAVING MIN(media_area) >= 5
        """, (trimestre, grupo_id))
        aprobados = len(cur.fetchall())

    pct_exito = round(aprobados * 100 / total_alumnos, 1) if total_alumnos > 0 else 0

    alumnos_suspensos = []
    if not es_infantil:
        cur.execute("""
            SELECT a.nombre, ar.nombre as area, AVG(e.nota) as media
            FROM evaluaciones e
            JOIN alumnos a ON e.alumno_id = a.id
            JOIN areas ar ON e.area_id = ar.id
            WHERE e.trimestre = ? AND a.grupo_id = ? AND a.deleted_at IS NULL
            GROUP BY e.alumno_id, e.area_id
            HAVING media < 5
            ORDER BY a.nombre, ar.nombre
        """, (trimestre, grupo_id))
        rows = cur.fetchall()
        alumno_map = {}
        for r in rows:
            if r["nombre"] not in alumno_map:
                alumno_map[r["nombre"]] = []
            alumno_map[r["nombre"]].append(f"{r['area']} ({r['media']:.1f})")
        alumnos_suspensos = [(n, ", ".join(a)) for n, a in alumno_map.items()]

    cur.execute("SELECT clave, valor FROM config WHERE clave LIKE 'logo_%' OR clave = 'tutor_firma_filename'")
    logo_config = {r["clave"]: r["valor"] for r in cur.fetchall()}

    from utils.db import get_app_data_dir
    uploads_dir = os.path.join(get_app_data_dir(), "uploads")

    def get_logo_path(lado):
        fn = logo_config.get(f"logo_{lado}_filename")
        if fn:
            p = os.path.join(uploads_dir, fn)
            if os.path.exists(p):
                return p
        return None

    logo_izda = get_logo_path("izda")
    logo_dcha = get_logo_path("dcha")
    pos_izda = logo_config.get("logo_izda_posicion", "left")
    pos_dcha = logo_config.get("logo_dcha_posicion", "right")

    firma_fn = logo_config.get("tutor_firma_filename")
    firma_path = None
    if firma_fn:
        p = os.path.join(uploads_dir, firma_fn)
        if os.path.exists(p):
            firma_path = p

    cur.execute("SELECT valor FROM config WHERE clave = 'nombre_centro'")
    r = cur.fetchone()
    nombre_centro = r["valor"] if r and r["valor"] else "CEIP"

    trim_texto = {"1": "PRIMERA", "2": "SEGUNDA", "3": "TERCERA"}.get(str(trimestre), trimestre.upper())

    firmantes = [l.strip() for l in equipo_informe.replace('\r', '').split('\n') if l.strip()]
    if not firmantes:
        firmantes = [l.strip() for l in equipo_docente_raw.replace('\r', '').split('\n') if l.strip()]

    from reportlab.platypus import Image as RLImage, HRFlowable
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=1.5*cm, bottomMargin=2*cm)
    elements = []
    styles = getSampleStyleSheet()

    style_title = ParagraphStyle('ActaTitle', parent=styles['Normal'],
                                fontSize=13, fontName='Helvetica-Bold',
                                alignment=1, spaceAfter=4)
    style_sub = ParagraphStyle('ActaSub', parent=styles['Normal'],
                                fontSize=10, fontName='Helvetica-Bold',
                                alignment=1, spaceAfter=8)
    style_body = ParagraphStyle('ActaBody', parent=styles['Normal'],
                                fontSize=10, leading=16, spaceAfter=6)
    style_label = ParagraphStyle('ActaLabel', parent=styles['Normal'],
                                fontSize=10, fontName='Helvetica-Bold', spaceAfter=4)

    logo_izda_el = None
    logo_dcha_el = None
    if logo_izda and os.path.exists(logo_izda):
        try:
            logo_izda_el = RLImage(logo_izda, width=3*cm, height=2*cm)
            logo_izda_el.hAlign = pos_izda.upper()
        except:
            logo_izda_el = None
    if logo_dcha and os.path.exists(logo_dcha):
        try:
            logo_dcha_el = RLImage(logo_dcha, width=3*cm, height=2*cm)
            logo_dcha_el.hAlign = pos_dcha.upper()
        except:
            logo_dcha_el = None

    from reportlab.platypus import Table as RLTable, TableStyle as RLTableStyle
    from reportlab.lib import colors as rl_colors

    col_izda = logo_izda_el if logo_izda_el else Paragraph(" ", styles['Normal'])
    cur.execute("SELECT valor FROM config WHERE clave = 'curso_escolar'")
    rc = cur.fetchone()
    curso_escolar = rc["valor"] if rc and rc["valor"] else ""
    col_centro = Paragraph(
        f"<b>{nombre_centro}</b><br/><b>ACTA DE LA {trim_texto} EVALUACIÓN</b><br/>Curso {curso_escolar}",
        ParagraphStyle('h', parent=styles['Normal'], alignment=1, fontSize=11, fontName='Helvetica-Bold', leading=16)
    )
    col_dcha = logo_dcha_el if logo_dcha_el else Paragraph(" ", styles['Normal'])

    header_tbl = RLTable([[col_izda, col_centro, col_dcha]], colWidths=[3.5*cm, 10*cm, 3.5*cm])
    header_tbl.setStyle(RLTableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (1, 0), (1, 0), 'CENTER'),
    ]))
    elements.append(header_tbl)
    elements.append(HRFlowable(width="100%", thickness=1, color=rl_colors.black, spaceAfter=10))

    hora_display = hora_str if hora_str else "……"
    fecha_display = fecha_str if fecha_str else "…………………………"
    lugar_fecha_str = f"<b>Lugar:</b> {nombre_centro}   |    <b>Fecha:</b> {fecha_display}   |    <b>Hora:</b> {hora_display}"
    elements.append(Paragraph(lugar_fecha_str, style_sub)) 
    elements.append(Spacer(1, 10))

    elements.append(Paragraph(
        f"Reunidos el Equipo Educativo de <b>{grupo_nombre}</b>, cuyos miembros asistentes se relacionan a continuación, "
        f"a fin de tratar el siguiente Orden del día:",
        style_body))
    elements.append(Spacer(1, 6))

    elements.append(Paragraph("<b>Asistentes:</b>", style_label))
    if firmantes:
        for f in firmantes:
            elements.append(Paragraph(f"• {f}", style_body))
    else:
        elements.append(Paragraph(".", style_body))
    elements.append(Spacer(1, 8))

    elements.append(Paragraph("<b>Orden del Día:</b>", style_label))
    orden = [
        "1. Valoración del Grupo.",
        "2. Dificultades encontradas en el grupo clase.",
        "3. Propuestas de mejora.",
        "4. Análisis de resultados de la evaluación.",
        "5. Alumnado con áreas suspensas" if not es_infantil else "5. Observaciones pedagógicas",
        "6. Conclusión y otras Observaciones.",
        "7. Miembros del Equipo Docente / Firmas.",
    ]
    for item in orden:
        elements.append(Paragraph(item, style_body))
    elements.append(Spacer(1, 10))
    elements.append(HRFlowable(width="100%", thickness=0.5, color=rl_colors.grey, spaceAfter=8))

    elements.append(Paragraph("<b>Desarrollo de la sesión:</b>", style_label))
    elements.append(Spacer(1, 4))

    elements.append(Paragraph("<b>1. Valoración del Grupo</b>", style_label))
    elements.append(Paragraph(valoracion if valoracion else "  ", style_body))
    elements.append(Spacer(1, 8))

    elements.append(Paragraph("<b>2. Dificultades encontradas en el grupo clase</b>", style_label))
    elements.append(Paragraph(dificultades if dificultades else "  ", style_body))
    elements.append(Spacer(1, 8))

    elements.append(Paragraph("<b>3. Propuestas de mejora</b>", style_label))
    elements.append(Paragraph(propuestas if propuestas else "  ", style_body))
    elements.append(Spacer(1, 8))

    elements.append(Paragraph("<b>4. Análisis de resultados de la evaluación</b>", style_label))
    elements.append(Paragraph(f"Número de alumnado: {total_alumnos}", style_body))
    elements.append(Paragraph(f"Alumnado con todo aprobado / promoción: {aprobados}", style_body))
    elements.append(Paragraph(f"Porcentaje de éxito: {pct_exito}%", style_body))
    elements.append(Spacer(1, 8))

    if not es_infantil and alumnos_suspensos:
        elements.append(Paragraph("<b>5. Alumnado con áreas suspensas</b>", style_label))
        data_susp = [["Alumno/a", "Áreas"]]
        for nombre_al, areas_al in alumnos_suspensos:
            data_susp.append([Paragraph(nombre_al, styles['Normal']), Paragraph(areas_al, styles['Normal'])])
        t_susp = RLTable(data_susp, colWidths=[7*cm, 10*cm])
        t_susp.setStyle(RLTableStyle([
            ('GRID', (0,0), (-1,-1), 0.5, rl_colors.grey),
            ('BACKGROUND', (0,0), (-1,0), rl_colors.lightgrey),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ]))
        elements.append(t_susp)
        elements.append(Spacer(1, 8))
    elif es_infantil:
        elements.append(Paragraph("<b>5. Observaciones pedagógicas</b>", style_label))
        elements.append(Paragraph("Ver informes individuales para detalle por alumno/a.", style_body))
        elements.append(Spacer(1, 8))
    else:
        elements.append(Paragraph("<b>5. Alumnado con áreas suspensas</b>", style_label))
        elements.append(Paragraph("✓ Todo el alumnado ha superado todas las áreas.", style_body))
        elements.append(Spacer(1, 8))

    elements.append(Paragraph("<b>6. Conclusión y otras Observaciones</b>", style_label))
    elements.append(Paragraph(otras_obs if otras_obs else "  ", style_body))
    elements.append(Spacer(1, 12))

    elements.append(Paragraph(f"En Valle de Guerra a {fecha_display}", style_body))
    elements.append(Spacer(1, 8))
    elements.append(Paragraph("<b>7. Miembros del Equipo Educativo / Firmas</b>", style_label))
    elements.append(Spacer(1, 6))

    if firmantes:
        sig_data = [["Docente", "Firma"]]
        for f in firmantes:
            row = [Paragraph(f, styles['Normal']), ""]
            # Si el firmante es el tutor, intentamos poner la firma
            # Se usa comparación flexible: busca coincidencia por palabras o si es el tutor
            es_tutor = False
            if firma_path and tutor_nombre:
                tutor_palabras = tutor_nombre.lower().split()
                firmante_lower = f.lower()
                # Si al menos 2 palabras del tutor están en el firmante, o el nombre completo está contenido
                if tutor_nombre.lower() in firmante_lower or \
                   (len(tutor_palabras) >= 2 and sum(1 for p in tutor_palabras if p in firmante_lower) >= 2) or \
                   ("tutor" in firmante_lower and "tutor" in tutor_nombre.lower()):
                    es_tutor = True
            
            if es_tutor:
                try:
                    img_f = RLImage(firma_path, width=4*cm, height=1.2*cm)
                    img_f.hAlign = 'CENTER'
                    row[1] = img_f
                except Exception as e:
                    print(f"Error insertando firma: {e}")
            sig_data.append(row)
        
        t_sig = RLTable(sig_data, colWidths=[9*cm, 8*cm],
                        rowHeights=[0.8*cm] + [1.5*cm]*len(firmantes))
        t_sig.setStyle(RLTableStyle([
            ('GRID', (0, 0), (-1,-1), 0.5, rl_colors.black),
            ('BACKGROUND', (0,0), (-1,0), rl_colors.lightgrey),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('ALIGN', (1, 1), (1, -1), 'CENTER'),
        ]))
        elements.append(t_sig)

    doc.build(elements)
    buffer.seek(0)

    filename = f"Acta_{trim_texto}_{grupo_nombre.replace(' ','_')}.pdf"
    return send_file(buffer, as_attachment=True, download_name=filename, mimetype='application/pdf')

@informes_bp.route("/api/informe/excel_grupo")
def excel_grupo():
    """Genera Excel con gráficas del grupo."""
    trimestre = request.args.get("trimestre")
    area_id = request.args.get("area_id")

    conn = get_db()
    cur = conn.cursor()

    grupo_id = session.get('active_group_id')

    cur.execute("SELECT nombre FROM grupos WHERE id = ?", (grupo_id,))
    row_g = cur.fetchone()
    grupo_nombre = row_g["nombre"] if row_g else "Grupo"

    cur.execute("""
        SELECT a.tipo_escala 
        FROM grupos g
        JOIN areas a ON g.etapa_id = a.etapa_id
        WHERE g.id = ? AND a.tipo_escala LIKE 'INFANTIL_%'
        LIMIT 1
    """, (grupo_id,))
    es_infantil = cur.fetchone() is not None

    periodo = f"T{trimestre}"

    if es_infantil:
        query_base = """
            SELECT e.alumno_id, e.area_id, e.nota
            FROM evaluaciones e
            WHERE e.trimestre = ?
            UNION ALL
            SELECT ec.alumno_id, c.area_id, ec.nota
            FROM evaluacion_criterios ec
            JOIN criterios c ON ec.criterio_id = c.id
            WHERE ec.periodo = ?
        """
    else:
        query_base = """
            SELECT e.alumno_id, e.area_id, e.nota
            FROM evaluaciones e
            WHERE e.trimestre = ? AND e.sda_id IS NOT NULL
        """

    if area_id:
        query_notas = f"""
        SELECT al.nombre as Alumno, ar.nombre as Area, AVG(combined.nota) as Nota, ar.tipo_escala
        FROM alumnos al
        JOIN areas ar ON ar.id = ?
        LEFT JOIN ({query_base}) combined ON al.id = combined.alumno_id AND ar.id = combined.area_id
        WHERE al.grupo_id = ?
        GROUP BY al.id, al.nombre, ar.id, ar.nombre, ar.tipo_escala
        ORDER BY al.nombre
        """
        if es_infantil:
            params_notas = [area_id, trimestre, periodo, grupo_id]
        else:
            params_notas = [area_id, trimestre, grupo_id]
    else:
        query_notas = f"""
        SELECT al.nombre as Alumno, ar.nombre as Area, AVG(combined.nota) as Nota, ar.tipo_escala
        FROM alumnos al
        CROSS JOIN (
            SELECT id, nombre, tipo_escala 
            FROM areas 
            WHERE etapa_id = (SELECT etapa_id FROM grupos WHERE id = ?)
        ) ar
        LEFT JOIN ({query_base}) combined ON al.id = combined.alumno_id AND ar.id = combined.area_id
        WHERE al.grupo_id = ?
        GROUP BY al.id, al.nombre, ar.id, ar.nombre, ar.tipo_escala
        ORDER BY al.nombre, ar.nombre
        """
        if es_infantil:
            params_notas = [grupo_id, trimestre, periodo, grupo_id]
        else:
            params_notas = [grupo_id, trimestre, grupo_id]

    df_notas = pd.read_sql_query(query_notas, conn, params=params_notas)

    if not df_notas.empty:
        def excel_format_nota(n, e):
            if pd.isna(n):
                return ""
            if (e and e.startswith("INFANTIL_")):
                rnd = round(n)
                if e == "INFANTIL_PA_A_MA":
                    return {1: "PA", 2: "AD", 3: "MA"}.get(rnd, "")
                return {1: "NI", 2: "EP", 3: "C"}.get(rnd, "")
            return n
        
        df_notas['Nota_Display'] = df_notas.apply(lambda r: excel_format_nota(r['Nota'], r['tipo_escala']), axis=1)

    cur.execute("SELECT * FROM informe_grupo WHERE trimestre = ? AND grupo_id = ?", (trimestre, grupo_id))
    row_grupo = cur.fetchone()

    grupo_data = {
        "Concepto": ["Valoración", "Propuestas", "Conclusión"],
        "Texto": [
            row_grupo["observaciones"] if row_grupo else "",
            row_grupo["propuestas_mejora"] if row_grupo else "",
            row_grupo["conclusion"] if row_grupo else ""
        ]
    }
    df_grupo = pd.DataFrame(grupo_data)

    query_susp = """
        SELECT a.nombre, COUNT(val.area_id) as Num_Suspensos
        FROM alumnos a
        LEFT JOIN (
            SELECT alumno_id, area_id, AVG(nota) as media_area
            FROM evaluaciones
            WHERE trimestre = ?
    """
    params_susp = [trimestre]
    if area_id:
        query_susp += " AND area_id = ?"
        params_susp.append(area_id)
    
    query_susp += """
            GROUP BY alumno_id, area_id
            HAVING media_area < 5
        ) val ON a.id = val.alumno_id
        WHERE a.grupo_id = ?
        GROUP BY a.id, a.nombre
        HAVING Num_Suspensos > 0
    """
    params_susp.append(grupo_id)

    cur.execute(query_susp, params_susp)
    susp_data = cur.fetchall()
    df_susp = pd.DataFrame(susp_data) if susp_data else pd.DataFrame(columns=["Alumno", "Num_Suspensos"])

    def detail_format_nota(row):
        n = row['Nota']
        e = row['tipo_escala']
        if pd.isna(n):
            return "—"
        if (e and e.startswith("INFANTIL_")):
            # Puede que n ya sea texto si viene de INFANTIL_PA_A_MA o INFANTIL_NI_EP_C original
            if isinstance(n, str):
                return n
            rnd = round(float(n))
            if e == "INFANTIL_PA_A_MA":
                return {1: "PA", 2: "AD", 3: "MA"}.get(rnd, "—")
            return {1: "NI", 2: "EP", 3: "C"}.get(rnd, "—")
        return f"{float(n):.2f}"

    if es_infantil:
        query_detalle = f"""
            SELECT al.nombre as Alumno, ar.nombre as Area, s.nombre as SDA, 
                   c.codigo as Criterio_Cod, c.descripcion as Criterio_Desc, 
                   e.nota as Nota, e.nivel as Nivel, ar.tipo_escala
            FROM alumnos al
            JOIN evaluaciones e ON al.id = e.alumno_id AND e.trimestre = ?
            JOIN areas ar ON e.area_id = ar.id
            JOIN criterios c ON e.criterio_id = c.id
            JOIN sda s ON e.sda_id = s.id
            WHERE al.grupo_id = ?
        """
        params_det = [trimestre, grupo_id]
        if area_id:
            query_detalle += " AND ar.id = ?"
            params_det.append(area_id)
            
        query_detalle += f"""
            UNION ALL
            SELECT al.nombre as Alumno, ar.nombre as Area, 'Criterios Directos' as SDA,
                   c.codigo as Criterio_Cod, c.descripcion as Criterio_Desc,
                   ec.nota as Nota, ec.nivel as Nivel, ar.tipo_escala
            FROM alumnos al
            JOIN evaluacion_criterios ec ON al.id = ec.alumno_id AND ec.periodo = ?
            JOIN criterios c ON ec.criterio_id = c.id
            JOIN areas ar ON c.area_id = ar.id
            WHERE al.grupo_id = ?
        """
        params_det.extend([periodo, grupo_id])
        if area_id:
            query_detalle += " AND ar.id = ?"
            params_det.append(area_id)
    else:
        query_detalle = f"""
            SELECT al.nombre as Alumno, ar.nombre as Area, s.nombre as SDA, 
                   c.codigo as Criterio_Cod, c.descripcion as Criterio_Desc, 
                   e.nota as Nota, e.nivel as Nivel, ar.tipo_escala
            FROM alumnos al
            JOIN evaluaciones e ON al.id = e.alumno_id AND e.trimestre = ?
            JOIN areas ar ON e.area_id = ar.id
            JOIN criterios c ON e.criterio_id = c.id
            JOIN sda s ON e.sda_id = s.id
            WHERE al.grupo_id = ?
        """
        params_det = [trimestre, grupo_id]
        if area_id:
            query_detalle += " AND ar.id = ?"
            params_det.append(area_id)

    query_detalle += " ORDER BY Alumno, Area, SDA, Criterio_Cod"

    df_detalle = pd.read_sql_query(query_detalle, conn, params=params_det)
    if not df_detalle.empty:
        df_detalle['Nota'] = df_detalle.apply(lambda r: detail_format_nota(r), axis=1)
        df_detalle = df_detalle.drop(columns=['tipo_escala'])

    cur.execute("SELECT count(*) FROM alumnos WHERE grupo_id = ?", (grupo_id,))
    total_alumnos = cur.fetchone()[0]

    susp_map = {0: 0, 1: 0, 2: 0, 3: 0}
    for r in susp_data:
        n = r["Num_Suspensos"]
        if n > 2:
            susp_map[3] += 1
        else:
            susp_map[n] += 1

    susp_map[0] = total_alumnos - len(susp_data)

    promo_stats = {
        "Categoría": ["Todo Aprobado", "1 Suspenso", "2 Suspensos", "+2 Suspensos"],
        "Cantidad": [susp_map[0], susp_map[1], susp_map[2], susp_map[3]],
        "Porcentaje": [
            f"{round(susp_map[0] * 100 / total_alumnos, 1)}%" if total_alumnos > 0 else "0%",
            f"{round(susp_map[1] * 100 / total_alumnos, 1)}%" if total_alumnos > 0 else "0%",
            f"{round(susp_map[2] * 100 / total_alumnos, 1)}%" if total_alumnos > 0 else "0%",
            f"{round(susp_map[3] * 100 / total_alumnos, 1)}%" if total_alumnos > 0 else "0%"
        ]
    }
    df_promo = pd.DataFrame(promo_stats)

    year = date.today().year
    if trimestre == "1":
        start_date, end_date = f"{year-1}-09-01", f"{year-1}-12-31"
    elif trimestre == "2":
        start_date, end_date = f"{year}-01-01", f"{year}-03-31"
    else:
        start_date, end_date = f"{year}-04-01", f"{year}-06-30"

    cur.execute("""
        SELECT a.nombre, estado, COUNT(*) as c
        FROM asistencia asi
        JOIN alumnos a ON a.id = asi.alumno_id
        WHERE asi.fecha BETWEEN ? AND ? AND a.grupo_id = ?
        GROUP BY a.id, a.nombre, estado
        ORDER BY c DESC
    """, (start_date, end_date, grupo_id))
    asist_rows = cur.fetchall()

    max_faltas = {"nombre": "Nadie", "num": 0}
    max_retrasos = {"nombre": "Nadie", "num": 0}

    temp_faltas = {}
    temp_retrasos = {}

    for r in asist_rows:
        if r["estado"] in ["falta_justificada", "falta_no_justificada"]:
            temp_faltas[r["nombre"]] = temp_faltas.get(r["nombre"], 0) + r["c"]
        elif r["estado"] == "retraso":
            temp_retrasos[r["nombre"]] = temp_retrasos.get(r["nombre"], 0) + r["c"]
        
    if temp_faltas:
        u_max = max(temp_faltas, key=temp_faltas.get)
        max_faltas = {"nombre": u_max, "num": temp_faltas[u_max]}
    
    if temp_retrasos:
        u_max = max(temp_retrasos, key=temp_retrasos.get)
        max_retrasos = {"nombre": u_max, "num": temp_retrasos[u_max]}

    asist_highlights = {
        "Tipo": ["Alumno con más faltas", "Alumno con más retrasos"],
        "Nombre": [max_faltas["nombre"], max_retrasos["nombre"]],
        "Cantidad": [max_faltas["num"], max_retrasos["num"]]
    }
    df_asist = pd.DataFrame(asist_highlights)

    fig1, ax1 = plt.subplots(figsize=(6, 4))
    colors_promo = ['#28a745', '#ffc107', '#fd7e14', '#dc3545']
    ax1.pie([susp_map[0], susp_map[1], susp_map[2], susp_map[3]], 
            labels=promo_stats["Categoría"],
            autopct='%1.1f%%',
            colors=colors_promo,
            startangle=90)
    ax1.set_title('Análisis de Promoción')
    promo_chart = BytesIO()
    plt.savefig(promo_chart, format='png', bbox_inches='tight')
    promo_chart.seek(0)
    plt.close()

    fig2, ax2 = plt.subplots(figsize=(6, 4))
    categories = ['Faltas', 'Retrasos']
    values = [max_faltas["num"], max_retrasos["num"]]
    colors_asist = ['#dc3545', '#fd7e14']
    ax2.bar(categories, values, color=colors_asist)
    ax2.set_ylabel('Cantidad')
    ax2.set_title('Asistencia Destacada (Máximos)')
    ax2.grid(axis='y', alpha=0.3)
    asist_chart = BytesIO()
    plt.savefig(asist_chart, format='png', bbox_inches='tight')
    asist_chart.seek(0)
    plt.close()

    if not df_notas.empty:
        area_avg = df_notas.groupby('Area')['Nota'].mean().sort_values(ascending=False)
        fig3, ax3 = plt.subplots(figsize=(8, 5))
        ax3.barh(area_avg.index, area_avg.values, color='#17a2b8')
        ax3.set_xlabel('Nota Media')
        ax3.set_title('Rendimiento por Área')
        ax3.grid(axis='x', alpha=0.3)
        area_chart = BytesIO()
        plt.savefig(area_chart, format='png', bbox_inches='tight')
        area_chart.seek(0)
        plt.close()
    else:
        area_chart = None

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        if not df_notas.empty:
            pivot = df_notas.pivot(index='Alumno', columns='Area', values='Nota_Display')
            pivot.to_excel(writer, sheet_name=f'Notas_T{trimestre}')
        
        df_grupo.to_excel(writer, sheet_name="Valoracion_Grupo", index=False)
        df_promo.to_excel(writer, sheet_name="Promocion", index=False, startrow=0)
        ws_promo = writer.sheets["Promocion"]
        
        from openpyxl.drawing.image import Image as XLImage
        img_promo = XLImage(promo_chart)
        ws_promo.add_image(img_promo, 'E2')
        
        if not df_susp.empty:
            df_susp.to_excel(writer, sheet_name="Alumnos_Suspensos", index=False)
        
        df_asist.to_excel(writer, sheet_name="Asistencia", index=False, startrow=0)
        ws_asist = writer.sheets["Asistencia"]
        
        img_asist = XLImage(asist_chart)
        ws_asist.add_image(img_asist, 'E2')

        if not df_detalle.empty:
            df_detalle.to_excel(writer, sheet_name="Detalle_Evaluacion", index=False)
        
        if area_chart:
            area_avg_df = area_avg.reset_index()
            area_avg_df.columns = ['Área', 'Nota Media']
            area_avg_df.to_excel(writer, sheet_name="Rendimiento_Areas", index=False)
            ws_areas = writer.sheets["Rendimiento_Areas"]
            img_areas = XLImage(area_chart)
            ws_areas.add_image(img_areas, 'D2')
    
    output.seek(0)
    filename = f"Informe_Grupo_{grupo_nombre.replace(' ', '_')}_T{trimestre}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@informes_bp.route("/api/informe/pdf_grupo")
def pdf_grupo():
    """PDF resumen del grupo."""
    trimestre = request.args.get("trimestre")
    colegio = request.args.get("colegio", "")
    fecha_inf = request.args.get("fecha_informe", "")
    area_id = request.args.get("area_id")
    
    conn = get_db()
    cur = conn.cursor()

    grupo_id = session.get('active_group_id')

    cur.execute("SELECT nombre FROM grupos WHERE id = ?", (grupo_id,))
    row_g = cur.fetchone()
    grupo_nombre = row_g["nombre"] if row_g else "Grupo"

    cur.execute("SELECT * FROM informe_grupo WHERE trimestre = ? AND grupo_id = ?", (trimestre, grupo_id))
    row_grupo = cur.fetchone()

    cur.execute("""
        SELECT a.tipo_escala 
        FROM grupos g
        JOIN areas a ON g.etapa_id = a.etapa_id
        WHERE g.id = ? AND a.tipo_escala LIKE 'INFANTIL_%'
        LIMIT 1
    """, (grupo_id,))
    es_infantil = cur.fetchone() is not None

    cur.execute("SELECT count(*) FROM alumnos WHERE grupo_id = ?", (grupo_id,))
    total_alumnos = cur.fetchone()[0]

    if total_alumnos <= 0:
        return "No hay alumnos en este grupo para generar el informe.", 400

    susp_map = {}
    infantil_map = {'C': 0, 'EP': 0, 'NI': 0}

    if es_infantil:
        periodo = f"T{trimestre}"
        cur.execute("""
            SELECT ec.alumno_id, ec.nota
            FROM evaluacion_criterios ec
            JOIN alumnos a ON ec.alumno_id = a.id
            WHERE ec.periodo = ? AND a.grupo_id = ?
        """, (periodo, grupo_id))
        notas_inf = cur.fetchall()
        
        alumnos_medias = {}
        for row in notas_inf:
            aid = row["alumno_id"]
            if aid not in alumnos_medias:
                alumnos_medias[aid] = []
            alumnos_medias[aid].append(row["nota"])
        
        evaluados_infantil = 0
        cur.execute("SELECT id FROM alumnos WHERE grupo_id = ?", (grupo_id,))
        for r in cur.fetchall():
            aid = r[0]
            if aid in alumnos_medias:
                evaluados_infantil += 1
                notas_validas = [n for n in alumnos_medias[aid] if n is not None]
                media = sum(notas_validas) / len(notas_validas) if notas_validas else 0
                
                # Escala dinámica
                cur.execute("""
                    SELECT a.tipo_escala FROM alumnos al 
                    JOIN grupos g ON al.grupo_id = g.id
                    JOIN areas a ON g.etapa_id = a.etapa_id
                    WHERE al.id = ? AND a.tipo_escala LIKE 'INFANTIL_%' LIMIT 1
                """, (aid,))
                row_esc = cur.fetchone()
                esc = row_esc["tipo_escala"] if row_esc else "INFANTIL_NI_EP_C"
                l1, l2, l3 = ("PA", "AD", "MA") if esc == "INFANTIL_PA_A_MA" else ("NI", "EP", "C")
                
                if media >= 2.5:
                    infantil_map[l3] = infantil_map.get(l3, 0) + 1
                elif media >= 1.5:
                    infantil_map[l2] = infantil_map.get(l2, 0) + 1
                else:
                    infantil_map[l1] = infantil_map.get(l1, 0) + 1
    else:
        cur.execute("""
            SELECT a.nombre, GROUP_CONCAT(ar.nombre || ' (' || ROUND(sub.media_area,1) || ')', ', ') as areas, COUNT(sub.area_id) as num_suspensos
            FROM alumnos a
            JOIN (
                SELECT alumno_id, area_id, AVG(nota) as media_area
                FROM evaluaciones
                WHERE trimestre = ?
                GROUP BY alumno_id, area_id
                HAVING media_area < 5
            ) sub ON a.id = sub.alumno_id
            JOIN areas ar ON sub.area_id = ar.id
            WHERE a.grupo_id = ? AND a.deleted_at IS NULL
            GROUP BY a.id, a.nombre
        """, (trimestre, grupo_id))
        suspensos_data = cur.fetchall()
    if not es_infantil and suspensos_data:
        susp_map = {0: 0, 1: 0, 2: 0, 3: 0}
        header_row = [Paragraph("<b>Resumen Promoción</b>", styles['Normal']), ""]
        for r in suspensos_data:
            n = r["num_suspensos"]
            if n > 2:
                susp_map[3] += 1
            else:
                susp_map[n] += 1
            
        susp_map[0] = total_alumnos - len(suspensos_data)

    year = date.today().year
    if trimestre == "1":
        start_date, end_date = f"{year-1}-09-01", f"{year-1}-12-31"
    elif trimestre == "2":
        start_date, end_date = f"{year}-01-01", f"{year}-03-31"
    else:
        start_date, end_date = f"{year}-04-01", f"{year}-06-30"

    cur.execute("""
        SELECT a.nombre, estado, COUNT(*) as c
        FROM asistencia asi
        JOIN alumnos a ON a.id = asi.alumno_id
        WHERE asi.fecha BETWEEN ? AND ? AND a.grupo_id = ?
        GROUP BY a.id, a.nombre, estado
        ORDER BY c DESC
    """, (start_date, end_date, grupo_id))
    asist_rows = cur.fetchall()

    max_faltas = {"nombre": "Nadie", "num": 0}
    max_retrasos = {"nombre": "Nadie", "num": 0}

    temp_faltas = {}
    temp_retrasos = {}

    for r in asist_rows:
        if r["estado"] in ["falta_justificada", "falta_no_justificada"]:
            temp_faltas[r["nombre"]] = temp_faltas.get(r["nombre"], 0) + r["c"]
        elif r["estado"] == "retraso":
            temp_retrasos[r["nombre"]] = temp_retrasos.get(r["nombre"], 0) + r["c"]
        
    if temp_faltas:
        u_max = max(temp_faltas, key=temp_faltas.get)
        max_faltas = {"nombre": u_max, "num": temp_faltas[u_max]}
    
    if temp_retrasos:
        u_max = max(temp_retrasos, key=temp_retrasos.get)
        max_retrasos = {"nombre": u_max, "num": temp_retrasos[u_max]}

    if es_infantil:
        cur.execute("""
            SELECT ar.nombre, AVG(val.nota) as media
            FROM (
                SELECT area_id, nota, alumno_id FROM evaluaciones WHERE trimestre = ?
                UNION ALL
                SELECT c.area_id, ec.nota, ec.alumno_id 
                FROM evaluacion_criterios ec
                JOIN criterios c ON ec.criterio_id = c.id
                WHERE ec.periodo = ?
            ) val
            JOIN areas ar ON val.area_id = ar.id
            JOIN alumnos al ON val.alumno_id = al.id
            WHERE al.grupo_id = ?
            GROUP BY ar.id, ar.nombre
            ORDER BY media DESC
        """, (trimestre, periodo, grupo_id))
    else:
        cur.execute("""
            SELECT ar.nombre, AVG(e.nota) as media
            FROM evaluaciones e
            JOIN areas ar ON e.area_id = ar.id
            JOIN alumnos al ON e.alumno_id = al.id
            WHERE e.trimestre = ? AND al.grupo_id = ? AND e.sda_id IS NOT NULL
            GROUP BY ar.id, ar.nombre
            ORDER BY media DESC
        """, (trimestre, grupo_id))
    area_perf_data = cur.fetchall()

    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    from reportlab.platypus import Image as RLImage

    fig1, ax1 = plt.subplots(figsize=(6, 4))
    ax1.set_title('Análisis de Promoción', pad=20)
    if es_infantil:
        cur.execute("""
            SELECT a.tipo_escala FROM grupos g
            JOIN areas a ON g.etapa_id = a.etapa_id
            WHERE g.id = ? AND a.tipo_escala LIKE 'INFANTIL_%' LIMIT 1
        """, (grupo_id,))
        esc_row = cur.fetchone()
        esc = esc_row["tipo_escala"] if esc_row else "INFANTIL_NI_EP_C"
        l1, l2, l3 = ("PA", "AD", "MA") if esc == "INFANTIL_PA_A_MA" else ("NI", "EP", "C")
        t1, t2, t3 = ("Parcialmente Adquirido", "Adquirido", "Muy Adquirido") if esc == "INFANTIL_PA_A_MA" else ("No Iniciado", "En proceso", "Conseguido")

        colors_promo = ['#28a745', '#ffc107', '#dc3545']
        labels_raw = [t3, t2, t1]
        data_raw = [infantil_map.get(l3, 0), infantil_map.get(l2, 0), infantil_map.get(l1, 0)]
        
        data, labels, colors_filtered = [], [], []
        for d, l, c in zip(data_raw, labels_raw, colors_promo):
            if d > 0:
                data.append(d)
                labels.append(l)
                colors_filtered.append(c)

        if sum(data) > 0:
            ax1.pie(data, labels=labels, autopct='%1.1f%%', pctdistance=0.8, labeldistance=1.1,
                    colors=colors_filtered, startangle=90, textprops={'fontsize': 9})
        else:
            ax1.text(0.5, 0.5, 'Sin datos para la gráfica', ha='center', va='center')
            ax1.axis('off')
    else:
        colors_promo = ['#28a745', '#ffc107', '#fd7e14', '#dc3545']
        labels_raw = ["Todo Aprobado", "1 Suspenso", "2 Suspensos", "+2 Suspensos"]
        data_raw = [susp_map[0], susp_map[1], susp_map[2], susp_map[3]]
        
        data, labels, colors_filtered = [], [], []
        for d, l, c in zip(data_raw, labels_raw, colors_promo):
            if d > 0:
                data.append(d)
                labels.append(l)
                colors_filtered.append(c)

        if sum(data) > 0:
            ax1.pie(data, labels=labels, autopct='%1.1f%%', pctdistance=0.8, labeldistance=1.1,
                    colors=colors_filtered, startangle=140, textprops={'fontsize': 9})
        else:
            ax1.text(0.5, 0.5, 'Sin datos para la gráfica', ha='center', va='center')
            ax1.axis('off')
    promo_chart_buf = BytesIO()
    plt.savefig(promo_chart_buf, format='png', bbox_inches='tight', dpi=300)
    promo_chart_buf.seek(0)
    plt.close()

    fig2, ax2 = plt.subplots(figsize=(6, 3.5))
    categories = ['Faltas\n' + max_faltas["nombre"], 'Retrasos\n' + max_retrasos["nombre"]]
    values = [max_faltas["num"], max_retrasos["num"]]
    colors_asist = ['#dc3545', '#fd7e14']
    ax2.bar(categories, values, color=colors_asist, width=0.6)
    ax2.set_ylabel('Cantidad')
    ax2.set_title('Asistencia Destacada')
    ax2.grid(axis='y', alpha=0.3)
    for i, v in enumerate(values):
        ax2.text(i, v + 0.5, str(v), ha='center', fontweight='bold')
    asist_chart_buf = BytesIO()
    plt.savefig(asist_chart_buf, format='png', bbox_inches='tight', dpi=300)
    asist_chart_buf.seek(0)
    plt.close()

    area_chart_buf = None
    if area_perf_data:
        fig3, ax3 = plt.subplots(figsize=(8, max(4, len(area_perf_data) * 0.5)))
        areas_n = [r["nombre"] for r in area_perf_data]
        medias_n = [r["media"] for r in area_perf_data]
        
        ax3.barh(areas_n, medias_n, color='#17a2b8')
        ax3.set_xlabel('Nota Media')
        ax3.set_title('Rendimiento Medio por Área')
        ax3.set_xlim(0, 10.5)
        ax3.grid(axis='x', alpha=0.3)
        
        for i, v in enumerate(medias_n):
            ax3.text(v + 0.1, i, f"{v:.2f}", va='center', fontweight='bold')
            
        area_chart_buf = BytesIO()
        plt.savefig(area_chart_buf, format='png', bbox_inches='tight', dpi=300)
        area_chart_buf.seek(0)
        plt.close()

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()

    add_header(elements, styles, f"Informe Global de Grupo - Trimestre {trimestre}")
    elements.append(Spacer(1, 20))

    if row_grupo:
        if row_grupo["observaciones"]:
            txt_formatted = row_grupo["observaciones"].replace('\n', '<br/>')
            elements.append(Paragraph("1. Valoración General del Tutor:", styles['Heading3']))
            elements.append(Paragraph(txt_formatted, styles['BodyText']))
            elements.append(Spacer(1, 10))
        
        dif_val = ""
        try:
            if "dificultades" in [k[0] for k in cur.description]:
                dif_val = row_grupo["dificultades"]
            elif "dificultades" in row_grupo.keys():
                dif_val = row_grupo["dificultades"]
        except:
            pass
        
        if dif_val:
            txt_formatted = dif_val.replace('\n', '<br/>')
            elements.append(Paragraph("2. Dificultades encontradas en el grupo clase:", styles['Heading3']))
            elements.append(Paragraph(txt_formatted, styles['BodyText']))
            elements.append(Spacer(1, 10))
        
        if row_grupo["propuestas_mejora"]:
            txt_formatted = row_grupo["propuestas_mejora"].replace('\n', '<br/>')
            elements.append(Paragraph("3. Propuestas de Mejora:", styles['Heading3']))
            elements.append(Paragraph(txt_formatted, styles['BodyText']))
            elements.append(Spacer(1, 10))
    else:
        elements.append(Paragraph("No hay valoración grupal registrada.", styles['Normal']))
    elements.append(Spacer(1, 15))

    elements.append(Paragraph("4. Análisis de resultados de la evaluación (Promoción/Rendimiento):", styles['Heading3']))

    def calc_pct(n):
        return f"{round(n * 100 / total_alumnos, 1)}%" if total_alumnos > 0 else "0%"

    if es_infantil:
        evaluados_infantil = sum(infantil_map.values())
        data_prom = [
            ["Conseguido (C)", f"{infantil_map['C']} ({round(infantil_map['C']*100/max(1,evaluados_infantil),1)}%)"],
            ["En Proceso (EP)", f"{infantil_map['EP']} ({round(infantil_map['EP']*100/max(1,evaluados_infantil),1)}%)"],
            ["No Iniciado (NI)", f"{infantil_map['NI']} ({round(infantil_map['NI']*100/max(1,evaluados_infantil),1)}%)"],
            ["No Evaluados", f"{total_alumnos - evaluados_infantil}"]
        ]
    else:
        data_prom = [
            ["Todo Aprobado", f"{susp_map[0]} ({calc_pct(susp_map[0])})"],
            ["1 Suspenso", f"{susp_map[1]} ({calc_pct(susp_map[1])})"],
            ["2 Suspensos", f"{susp_map[2]} ({calc_pct(susp_map[2])})"],
            ["+2 Suspensos", f"{susp_map[3]} ({calc_pct(susp_map[3])})"]
        ]
    t_prom = Table(data_prom, colWidths=[4*cm, 4*cm])
    t_prom.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 1, colors.grey),
        ('BACKGROUND', (0,0), (0,-1), colors.lightgrey)
    ]))
    elements.append(t_prom)
    elements.append(Spacer(1, 10))

    promo_img = RLImage(promo_chart_buf, width=12*cm, height=8*cm)
    elements.append(promo_img)
    elements.append(Spacer(1, 15))

    elements.append(Paragraph("Datos de Asistencia Destacados:", styles['Heading3']))
    elements.append(Paragraph(f"• Alumno/a con más faltas: <b>{max_faltas['nombre']}</b> ({max_faltas['num']} faltas)", styles['Normal']))
    elements.append(Paragraph(f"• Alumno/a con más retrasos: <b>{max_retrasos['nombre']}</b> ({max_retrasos['num']} retrasos)", styles['Normal']))
    elements.append(Spacer(1, 10))

    asist_img = RLImage(asist_chart_buf, width=12*cm, height=7*cm)
    elements.append(asist_img)
    elements.append(Spacer(1, 15))

    if area_perf_data and area_chart_buf:
        elements.append(PageBreak())
        add_header(elements, styles, f"Rendimiento por Áreas - Trimestre {trimestre}")
        elements.append(Paragraph("Resumen de Calificaciones Medias por Área:", styles['Heading3']))
        
        data_area = [["Área", "Nota Media"]]
        for r in area_perf_data:
            data_area.append([r["nombre"], f"{r['media']:.2f}"])
            
        t_area = Table(data_area, colWidths=[10*cm, 4*cm])
        t_area.setStyle(TableStyle([
            ('GRID', (0,0), (-1,-1), 1, colors.grey),
            ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
            ('ALIGN', (1,0), (1,-1), 'CENTER')
        ]))
        elements.append(t_area)
        elements.append(Spacer(1, 15))
        
        area_img = RLImage(area_chart_buf, width=16*cm, height=max(8, len(area_perf_data)*0.8)*cm)
        elements.append(area_img)
        
    if not es_infantil and suspensos_data:
        elements.append(Paragraph("5. Alumnado con áreas suspensas:", styles['Heading3']))
        data_susp = [["Alumno/a", "Áreas"]]
        for r in suspensos_data:            data_susp.append([Paragraph(r["nombre"], styles['Normal']), Paragraph(r["areas"], styles['Normal'])])
        t_susp = Table(data_susp, colWidths=[6*cm, 10*cm])
        t_susp.setStyle(TableStyle([
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
            ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ]))
        elements.append(t_susp)
        elements.append(Spacer(1, 15))
    elif es_infantil:
        elements.append(Paragraph("5. Observaciones pedagógicas:", styles['Heading3']))
        elements.append(Paragraph("Ver informes individuales para detalle por alumno/a.", styles['Normal']))
        elements.append(Spacer(1, 15))

    if row_grupo and row_grupo["conclusion"]:
        txt_formatted = row_grupo["conclusion"].replace('\n', '<br/>')
        elements.append(Paragraph("6. Conclusión Final / Otras Observaciones:", styles['Heading3']))
        elements.append(Paragraph(txt_formatted, styles['BodyText']))
        elements.append(Spacer(1, 15))

    equipo_docente = ""
    if row_grupo:
        try:
            equipo_docente = row_grupo["equipo_docente"] if "equipo_docente" in [k[0] for k in cur.description] else ""
        except:
            equipo_docente = ""
        
    if equipo_docente:
        elements.append(Spacer(1, 10))
        elements.append(Paragraph("7. Miembros del Equipo Docente (Firmas)", styles['Heading3']))
        
        firmantes = [line.strip() for line in equipo_docente.replace('\r', '').split('\n') if line.strip()]
        if firmantes:
            sig_data = [["Docente", "Firma"]]
            for f in firmantes:
                sig_data.append([Paragraph(f, styles['Normal']), ""])
                
            t_sig = Table(sig_data, colWidths=[8*cm, 8*cm], rowHeights=[1*cm] + [1.5*cm]*len(firmantes))
            t_sig.setStyle(TableStyle([
                ('GRID', (0,0), (-1,-1), 1, colors.black),
                ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('ALIGN', (0,0), (-1,0), 'CENTER'),
            ]))
            elements.append(t_sig)

    doc.build(elements)
    buffer.seek(0)

    filename = f"Informe_Grupo_{grupo_nombre.replace(' ', '_')}_T{trimestre}.pdf"

    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/pdf'
    )

@informes_bp.route("/api/rubricas/pdf/<int:sda_id>")
def rubrica_pdf_sda(sda_id):
    conn = get_db()
    cur = conn.cursor()
    
    cur.execute("SELECT nombre FROM sda WHERE id = ?", (sda_id,))
    row = cur.fetchone()
    if not row:
        return "SDA no encontrada", 404
    sda_name = row["nombre"]

    cur.execute("""
        SELECT c.id, c.codigo, c.descripcion
        FROM criterios c
        JOIN sda_criterios sc ON sc.criterio_id = c.id
        WHERE sc.sda_id = ?
        ORDER BY c.id
    """, (sda_id,))
    criterios = cur.fetchall()

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()

    elements.append(Paragraph(f"Rúbricas: {sda_name}", styles['Title']))
    elements.append(Spacer(1, 12))

    for c in criterios:
        elements.append(Paragraph(f"<b>Criterio {c['codigo']}:</b>", styles['Heading3']))
        elements.append(Paragraph(c['descripcion'], styles['Normal']))
        elements.append(Spacer(1, 8))
        
        cur.execute("SELECT nivel, descriptor FROM rubricas WHERE criterio_id = ? ORDER BY nivel", (c['id'],))
        rubs = cur.fetchall()
        
        if rubs:
            data = [["Nivel", "Descriptor"]]
            for r in rubs:
                data.append([str(r['nivel']), Paragraph(r['descriptor'], styles['Normal'])])
            
            t = Table(data, colWidths=[2*cm, 14*cm])
            t.setStyle(TableStyle([
                ('GRID', (0,0), (-1,-1), 1, colors.black),
                ('VALIGN', (0,0), (-1,-1), 'TOP'),
                ('BACKGROUND', (0,0), (-1,0), colors.lightgrey)
            ]))
            elements.append(t)
        else:
            elements.append(Paragraph("<i>No hay rúbricas definidas para este criterio.</i>", styles['Normal']))
        
        elements.append(Spacer(1, 15))
    
    doc.build(elements)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"Rubricas_{sda_name}.pdf",
        mimetype='application/pdf'
    )